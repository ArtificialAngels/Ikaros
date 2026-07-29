"""
吉利 GEELY BHE20-PHEV 模具 2D 文件清单
项目号：25-8716
特殊处理：
  1. 范围编号（如 7220-7222、6002-6009）
  2. "（已打印）/(已打印)" 后缀
  3. 连续数字段（3002-3007_动模镶件）
"""
import re
from pathlib import Path
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r"E:\KPSNC模具资料\GEELY_BHE20_PHEV\2026-04-08_25-8716_Geely BHE20-PHEV缸体压铸模_交货资料\1.2D、3D图纸数据\2D")
OUTPUT = ROOT / "模具文件清单.xlsx"

# === 解析规则 ===
PROJECT_PREFIXES = [r"25-8716"]
PREFIX_PATTERN = re.compile(r"^(" + "|".join(PROJECT_PREFIXES) + r")[\s_\-]*", re.IGNORECASE)

# 单个编号或范围编号：可在 _ 或字符串开头/单词边界
SINGLE_NUM = re.compile(r"(?:^|_)(\d{4,5})$")
RANGE_NUM = re.compile(r"(?:^|_)(\d{4,5})\s*[-~]\s*(\d{4,5})$")
# 范围+描述混合
INNER_RANGE = re.compile(r"(?:^|_)(\d{4,5})\s*[-~]\s*(\d{4,5})")
INNER_SINGLE = re.compile(r"(?:^|_)(\d{4,5})")
HAS_CHINESE = re.compile(r"[\u4e00-\u9fff]")

# 已打印后缀（全角/半角括号、空白）
PRINTED_SUFFIX = re.compile(r"[（(]\s*已打印\s*[)）]\s*$")


def parse_filename(stem: str) -> dict:
    original = stem
    s = PREFIX_PATTERN.sub("", stem, count=1)
    prefix_removed = s

    # ★ 关键修复：先去"已打印"后缀，再匹配编号
    s = PRINTED_SUFFIX.sub("", s).strip()

    # 优先级策略：
    #   A) 末尾范围编号 (7121-7127)  - 文件名最明确
    #   B) 末尾单编号  (7121)         - 文件名最明确
    #   C) 开头第一个 4-5 位编号      - 如 7452_S4滑块...，避免取中间"1_4"
    #   D) 任意位置范围编号
    #   E) 任意位置单编号

    range_start = ""
    range_end = ""

    m = RANGE_NUM.search(s)
    if m:
        range_start, range_end = m.group(1), m.group(2)
        desc_part = s[:m.start()]
    else:
        m2 = SINGLE_NUM.search(s)
        if m2:
            range_start = range_end = m2.group(1)
            desc_part = s[:m2.start()]
        else:
            # 取开头第一个 4-5 位编号（避免中间干扰）
            lead = re.search(r"^(\d{4,5})(?:_|$)", s)
            if lead:
                range_start = range_end = lead.group(1)
                desc_part = s[lead.end():]
            else:
                m3 = INNER_RANGE.search(s)
                if m3:
                    range_start, range_end = m3.group(1), m3.group(2)
                    desc_part = s.replace(m3.group(0), "")
                else:
                    m4 = INNER_SINGLE.search(s)
                    if m4:
                        range_start = range_end = m4.group(1)
                        desc_part = s.replace(m4.group(0), "")
                    else:
                        desc_part = s

    # 清理尾部下划线/连字符
    desc_part = re.sub(r"[\s_\-]+$", "", desc_part)

    # 合并所有含中文的 _ 分段（避免被切碎成短段，保留"型号+描述"完整信息）
    if HAS_CHINESE.search(desc_part):
        parts = desc_part.split("_")
        chinese_parts = [p for p in parts if HAS_CHINESE.search(p)]
        if chinese_parts:
            description = "".join(chinese_parts)
        else:
            description = desc_part
    else:
        description = desc_part

    # 规范化：
    #   1) 切分时丢失的分数下划线 (1_4、1_8、3_8) 还原为 "/"
    #     "滑块14转M16水管" -> "滑块1/4转M16水管"  (合并后"14"实际是"1"+"4")
    #     "滑块18水管" -> "滑块1/8水管"
    #     "滑块38水管" -> "滑块3/8水管"
    #     规则：单数字 + 数字 + 紧跟中文/字母 -> 解释为分数
    description = re.sub(r"(?<=[滑块])([1-9])(?=[1-9][一-鿿A-Za-z])", r"\1/", description)
    description = description.strip()

    # 编号展示
    if range_start and range_end and range_start != range_end:
        number_display = f"{range_start}-{range_end}"
        number_count = int(range_end) - int(range_start) + 1
    elif range_start:
        number_display = range_start
        number_count = 1
    else:
        number_display = ""
        number_count = 0

    return {
        "original": original,
        "prefix_removed": prefix_removed,
        "description": description,
        "number": number_display,
        "number_start": range_start,
        "number_end": range_end,
        "number_count": number_count,
    }


def main():
    print(f"扫描根目录: {ROOT}")
    if not ROOT.exists():
        raise SystemExit(f"❌ 目录不存在: {ROOT}")

    rows = []
    subdirs = sorted([p for p in ROOT.iterdir() if p.is_dir()])
    print(f"发现 {len(subdirs)} 个子目录")

    total = 0
    for sd in subdirs:
        files = sorted([p for p in sd.iterdir() if p.is_file()])
        print(f"  📁 {sd.name}: {len(files)} 个文件")
        for fp in files:
            total += 1
            stem = fp.stem
            ext = fp.suffix.lower().lstrip(".")

            p = parse_filename(stem)
            remarks = []
            if p["number_count"] > 1:
                remarks.append(f"范围编号({p['number_count']}件)")
            if "已打印" in stem:
                remarks.append("已打印")
            if "散件" in stem or "结构件" in stem:
                remarks.append("散件/结构件")
            if not p["number"]:
                remarks.append("无编号")

            rows.append({
                "序号": total,
                "子目录": sd.name,
                "原文件名": fp.name,
                "中文描述": p["description"],
                "编号": p["number"],
                "零件数量": p["number_count"],
                "扩展名": ext,
                "备注": "; ".join(remarks),
            })

    print(f"\n共解析 {total} 个文件")

    # ============ 写 Excel ============
    wb = Workbook()
    ws = wb.active
    ws.title = "模具文件清单"

    headers = ["序号", "子目录", "原文件名", "中文描述", "编号", "零件数量", "扩展名", "备注"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="305496")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="999999")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    data_align = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for col_idx, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border

    for r in rows:
        ws.append([r["序号"], r["子目录"], r["原文件名"], r["中文描述"],
                   r["编号"], r["零件数量"], r["扩展名"], r["备注"]])
        row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = data_align
            cell.border = border
        # 范围编号行加浅绿
        if r["零件数量"] > 1:
            green = PatternFill("solid", fgColor="E2EFDA")
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = green
        # 散件/结构件行加浅黄
        elif "散件/结构件" in r["备注"]:
            yellow = PatternFill("solid", fgColor="FFF2CC")
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = yellow

    col_widths = {1: 6, 2: 28, 3: 60, 4: 38, 5: 14, 6: 10, 7: 8, 8: 22}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ============ 子目录统计 Sheet ============
    ws2 = wb.create_sheet("子目录统计")
    ws2.append(["子目录", "文件数", "覆盖零件数(去重)", "占总数%"])
    for col_idx in range(1, 5):
        c = ws2.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border

    sub_counter = Counter(r["子目录"] for r in rows)
    sub_parts = Counter()
    for r in rows:
        if r["零件数量"] == 1:
            sub_parts[r["子目录"]] += 1
        else:
            sub_parts[r["子目录"]] += r["零件数量"]

    total_n = len(rows)
    total_parts = sum(r["零件数量"] for r in rows)
    for sub, cnt in sub_counter.most_common():
        parts = sub_parts[sub]
        pct = cnt * 100 / total_n
        ws2.append([sub, cnt, parts, f"{pct:.1f}%"])
        for col_idx in range(1, 5):
            cell = ws2.cell(row=ws2.max_row, column=col_idx)
            cell.border = border
            cell.alignment = data_align

    ws2.append([])
    ws2.append(["合计", total_n, total_parts, "100.0%"])
    for col_idx in range(1, 5):
        cell = ws2.cell(row=ws2.max_row, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.border = border
    for col, w in [(1, 30), (2, 10), (3, 18), (4, 12)]:
        ws2.column_dimensions[get_column_letter(col)].width = w

    wb.save(OUTPUT)
    print(f"\n✅ Excel 已保存: {OUTPUT}")
    print(f"   文件大小: {OUTPUT.stat().st_size / 1024:.1f} KB")

    # 抽检
    print("\n=== 抽检前 12 行 ===")
    print("-" * 120)
    for r in rows[:12]:
        print(f"  [{r['序号']:>2}] {r['子目录']:<25} | 描述: {r['中文描述']:<28} | 编号: {r['编号']:<10} | 数量: {r['零件数量']}")


if __name__ == "__main__":
    main()
