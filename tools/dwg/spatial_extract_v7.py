"""
完整提取多图框表格 - v7 终极版
策略：
1. 找所有"件号"位置作为表格的"件号列 X 锚点"
2. 同一表格的特征：件号 X 相同，向下 30-100mm 内出现数字数据
3. 找"高压点冷"作为分组标题
4. 直接按 Y 聚类所有 AM_4 文字为行，再按列 X 范围匹配
5. 标题行（同时含 5 关键词的 Y 簇）作为列定义
6. 数据行：所有"非标题"的 AM_4 文字
"""
import ezdxf
from pathlib import Path
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DXF_PATH = Path(r'E:\KPSNC模具资料\GEELY_BHE20_ICE\2026-04-08_25-8715_Geely BHE20-ICE\1.2D3DMOLd\2D\COOLING\25-8715_点冷却器（已打印）.dxf')
OUTPUT_XLSX = DXF_PATH.parent / "DXF表格提取.xlsx"

def clean_mtext(text):
    text = re.sub(r'\{[^}]*\}', '', text)
    text = re.sub(r'\\[WwFfTtCcPpQqAaLlKk][^;]*;', '', text)
    text = re.sub(r'\\[WwFfTtCcPpQqAaLlKk]', '', text)
    text = text.replace('%%c', 'Ø')
    return re.sub(r'\s+', ' ', text).strip()

print(f"读取: {DXF_PATH.name}")
doc = ezdxf.readfile(str(DXF_PATH))
msp = doc.modelspace()
entities = list(msp)

# 收集文字
text_items = []
for e in entities:
    if e.dxftype() in ('TEXT', 'MTEXT'):
        try:
            raw = e.dxf.text if e.dxftype() == 'TEXT' else e.text
            if not raw or not raw.strip():
                continue
            clean = clean_mtext(raw)
            if not clean:
                continue
            pos = e.dxf.insert
            text_items.append({
                'x': float(pos[0]), 'y': float(pos[1]),
                'text': clean, 'raw': raw, 'layer': e.dxf.layer,
            })
        except:
            pass
print(f"文字总数: {len(text_items)}")

# ===== 第一步：找所有 5 关键词都在的 Y 簇（标题行） =====
sorted_by_y = sorted(text_items, key=lambda t: -t['y'])
y_groups = []
current = [sorted_by_y[0]]
for t in sorted_by_y[1:]:
    if abs(t['y'] - current[-1]['y']) < 2.0:
        current.append(t)
    else:
        y_groups.append(current)
        current = [t]
y_groups.append(current)

# 找标题行：含 5 关键词且都在 AM_4
HEADER_KEYS = {'件号', '冷却编号', '型号', '使用零件', '数量'}
COLUMN_NAMES = ['件号', '冷却编号', '型号', 'd', 'L', 'E', 'J', '使用零件', '运水孔K', '数量']
header_ys = []  # [(y, header_texts), ...]
for grp in y_groups:
    texts = [t['text'] for t in grp]
    if all(k in texts for k in HEADER_KEYS) and all(t['layer'] == 'AM_4' for t in grp):
        # 这个 Y 簇是标题行
        header_ys.append({'y': grp[0]['y'], 'texts': sorted(grp, key=lambda t: t['x'])})

print(f"找到 {len(header_ys)} 个标题行")
for h in header_ys:
    print(f"  Y={h['y']:.1f}: {len(h['texts'])} 个文字")

# ===== 第二步：识别"件号"列的所有不同 X =====
# "件号" X 不同的 = 不同表格
# 但同一"表格"也可能有多行（5 数据行）
# 一个表格 = 一个"件号" X 锚点
jianhao_xs = set()
for h in header_ys:
    for t in h['texts']:
        if t['text'] == '件号':
            jianhao_xs.add(round(t['x'], 1))

print(f"\n'件号' X 锚点: {len(jianhao_xs)} 个")
for x in sorted(jianhao_xs):
    print(f"  X={x}")

# ===== 第三步：对每个"件号" X，找离它最近的标题行 + 数据 =====
# 但同一"件号" X 可能有多个标题 Y（不同图框的相同 X 锚点）
# 关键：找最近的标题行（在该 X 锚点下方最近 50mm 范围内）

tables = []
for jx in sorted(jianhao_xs):
    # 找该 X 附近的所有标题行（X 差 < 5mm）
    related_headers = [h for h in header_ys if any(abs(t['x'] - jx) < 5 for t in h['texts'] if t['text'] == '件号')]
    if not related_headers:
        continue

    # 对每个相关标题行，提取一个表格
    for h in related_headers:
        a_y = h['y']
        a_x = jx

        # 列 X 中心
        col_x_map = {}
        for i, t in enumerate(h['texts']):
            if i < len(COLUMN_NAMES):
                col_x_map[COLUMN_NAMES[i]] = t['x']

        # 数据行：a_x ± 200, Y < a_y - 3
        # Y 范围：到下一个标题行 Y（30mm 以上）
        next_header_y = None
        for h2 in header_ys:
            if h2['y'] < a_y - 30:  # 至少低于 30mm
                if next_header_y is None or h2['y'] > next_header_y:
                    next_header_y = h2['y']
        y_bottom = next_header_y if next_header_y else a_y - 200

        data_texts = [t for t in text_items
                      if abs(t['x'] - a_x) < 250
                      and y_bottom <= t['y'] <= a_y - 3
                      and t['layer'] == 'AM_4'
                      and t['text'] not in COLUMN_NAMES]

        # Y 聚类
        sorted_data = sorted(data_texts, key=lambda t: -t['y'])
        rows = []
        if sorted_data:
            current = [sorted_data[0]]
            for t in sorted_data[1:]:
                if abs(t['y'] - current[-1]['y']) < 3.0:
                    current.append(t)
                else:
                    current.sort(key=lambda t: t['x'])
                    rows.append(current)
                    current = [t]
            current.sort(key=lambda t: t['x'])
            rows.append(current)

        tables.append({
            'anchor_x': a_x, 'anchor_y': a_y,
            'col_x_map': col_x_map,
            'data_rows': rows,
        })

print(f"\n共识别 {len(tables)} 个表格")
for i, t in enumerate(tables):
    print(f"  表 {i+1}: 锚点 ({t['anchor_x']:.0f}, {t['anchor_y']:.0f}), 数据 {len(t['data_rows'])} 行")

# ===== 输出 Excel =====
wb = Workbook()
wb.remove(wb.active)

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="305496")
thin = Side(border_style="thin", color="999999")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
data_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center")

def assign_to_col(t_x, col_x_map, threshold=20):
    best_col = None
    best_dist = threshold
    for col, x in col_x_map.items():
        d = abs(t_x - x)
        if d < best_dist:
            best_dist = d
            best_col = col
    return best_col

# Sheet 1: 全部表格汇总
ws = wb.create_sheet("全部表格汇总")
ws.append(['表#', '锚点 X', '锚点 Y', '行号'] + COLUMN_NAMES)
for c in ws[1]:
    c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = border

for t_idx, tbl in enumerate(tables):
    for r_idx, row in enumerate(tbl['data_rows']):
        row_data = {c: '' for c in COLUMN_NAMES}
        for t in row:
            col = assign_to_col(t['x'], tbl['col_x_map'])
            if col and t['text']:
                if row_data[col]:
                    row_data[col] += ' ' + t['text']
                else:
                    row_data[col] = t['text']
        ws.append([t_idx + 1, round(tbl['anchor_x']), round(tbl['anchor_y']), r_idx + 1] +
                  [row_data[c] for c in COLUMN_NAMES])
        r = ws.max_row
        for c_idx in range(1, 15):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = border; cell.alignment = data_align

for col, w in [('A', 6), ('B', 10), ('C', 10), ('D', 6),
               ('E', 25), ('F', 25), ('G', 15), ('H', 12), ('I', 8),
               ('J', 8), ('K', 8), ('L', 25), ('M', 15), ('N', 8)]:
    ws.column_dimensions[col].width = w
ws.freeze_panes = "E2"
ws.auto_filter.ref = ws.dimensions

# 每个表格单独 Sheet
for t_idx, tbl in enumerate(tables):
    sheet_name = f"表{t_idx+1}"
    if len(sheet_name) > 31:
        sheet_name = f"T{t_idx+1}"
    ws = wb.create_sheet(sheet_name)
    ws.append(COLUMN_NAMES)
    for c in ws[1]:
        c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = border

    for row in tbl['data_rows']:
        row_data = {c: '' for c in COLUMN_NAMES}
        for t in row:
            col = assign_to_col(t['x'], tbl['col_x_map'])
            if col and t['text']:
                if row_data[col]:
                    row_data[col] += ' ' + t['text']
                else:
                    row_data[col] = t['text']
        ws.append([row_data[c] for c in COLUMN_NAMES])
        r = ws.max_row
        for c_idx in range(1, 11):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = border; cell.alignment = data_align

    for col, w in [('A', 25), ('B', 25), ('C', 15), ('D', 12), ('E', 8),
                  ('F', 8), ('G', 8), ('H', 25), ('I', 15), ('J', 8)]:
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

wb.save(OUTPUT_XLSX)
print(f"\n✅ Excel 已保存: {OUTPUT_XLSX}")
print(f"   大小: {OUTPUT_XLSX.stat().st_size/1024:.1f} KB")
print(f"\nSheet 列表:")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"  {name}: {ws.max_row-1} 行 x {ws.max_column} 列")
