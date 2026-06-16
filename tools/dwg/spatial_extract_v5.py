"""
完整提取多图框表格 - v5 最终版
策略：
1. 找"件号"作为图框锚点（AM_4 图层）
2. 每个图框内的标题行：取锚点 Y ± 2mm 的所有 AM_4 文字
3. 数据行：标题 Y - 5mm 到 Y - 300mm 范围
4. 用标题行每个文字的 X 作为"列中心"，数据按 X 距离归类（阈值 20mm）
"""
import ezdxf
from pathlib import Path
from collections import Counter, defaultdict
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DXF_PATH = Path(r'E:\KPSNC模具资料\GEELY_BHE20_ICE\2026-04-08_25-8715_Geely BHE20-ICE\1.2D3DMOLd\2D\COOLING\25-8715_点冷却器（已打印）.dxf')
OUTPUT_XLSX = DXF_PATH.parent / "DXF表格提取.xlsx"

def clean_mtext(text: str) -> str:
    text = re.sub(r'\{[^}]*\}', '', text)
    text = re.sub(r'\\[WwFfTtCcPpQqAaLlKk][^;]*;', '', text)
    text = re.sub(r'\\[WwFfTtCcPpQqAaLlKk]', '', text)
    text = text.replace('%%c', 'Ø')
    text = re.sub(r'\s+', ' ', text).strip()
    return text

print(f"读取: {DXF_PATH.name}")
doc = ezdxf.readfile(str(DXF_PATH))
msp = doc.modelspace()
entities = list(msp)

# 收集所有文字
text_items = []
for e in entities:
    if e.dxftype() in ('TEXT', 'MTEXT'):
        try:
            raw = e.dxf.text if e.dxftype() == 'TEXT' else e.text
            if not raw or not raw.strip():
                continue
            clean = clean_mtext(raw)
            if not clean:
                continue
            pos = e.dxf.insert
            text_items.append({
                'x': float(pos[0]),
                'y': float(pos[1]),
                'text': clean,
                'raw': raw,
                'layer': e.dxf.layer,
            })
        except:
            pass

print(f"文字总数: {len(text_items)}")

# ===== 第一步：找所有"件号"位置（AM_4）=====
# 一个图框可能没有"件号"，但只要标题行有 5 个核心词就算
# 找同时含 5 关键词的 Y 簇
sorted_by_y = sorted(text_items, key=lambda t: -t['y'])
y_groups = []
current = [sorted_by_y[0]]
for t in sorted_by_y[1:]:
    if abs(t['y'] - current[-1]['y']) < 2.0:
        current.append(t)
    else:
        y_groups.append(current)
        current = [t]
y_groups.append(current)

# 找标题行（同时含 5 关键词）
header_y_set = set()
for grp in y_groups:
    texts = [t['text'] for t in grp]
    if all(k in texts for k in ['件号', '冷却编号', '型号', '使用零件', '数量']):
        if all(t['layer'] == 'AM_4' for t in grp):
            header_y_set.add(round(grp[0]['y'], 2))

print(f"找到 {len(header_y_set)} 个标题行 Y")
header_ys = sorted(header_y_set, reverse=True)
print(f"Y 范围: {min(header_ys):.0f} ~ {max(header_ys):.0f}")

# ===== 第二步：合并相邻标题行为图框组 =====
# 一个图框里有多套标题（同一图框的多个小表），Y 距离 < 30mm 视为同图框
TABLE_GAP = 30  # mm
merged = []  # [[y1, y2, ...], ...]
if header_ys:
    current_group = [header_ys[0]]
    for y in header_ys[1:]:
        if current_group[-1] - y < TABLE_GAP:
            current_group.append(y)
        else:
            merged.append(current_group)
            current_group = [y]
    merged.append(current_group)

print(f"\n合并后图框组: {len(merged)}")
for i, g in enumerate(merged):
    print(f"  图框 {i+1}: Y={min(g):.0f} ~ {max(g):.0f}, {len(g)} 个标题行")

# ===== 第三步：对每个图框组，提取标题行文字+数据行 =====
COLUMN_NAMES = ['件号', '冷却编号', '型号', 'd', 'L', 'E', 'J', '使用零件', '运水孔K', '数量']

tables = []
for grp in merged:
    y_top = max(grp) + 2  # 标题行最上 Y
    y_bottom = min(grp) - 300  # 表格底部 Y（数据行最多向下 300mm）

    # 该图框的标题行文字
    header_texts = []
    for y in grp:
        for t in text_items:
            if abs(t['y'] - y) < 2.0 and t['layer'] == 'AM_4' and t['text'] in COLUMN_NAMES:
                header_texts.append(t)
    header_texts.sort(key=lambda t: t['x'])

    # 列 X 中心
    col_x_map = {}
    for i, t in enumerate(header_texts):
        if i < len(COLUMN_NAMES):
            col_x_map[COLUMN_NAMES[i]] = t['x']

    # 数据行
    data_texts = [t for t in text_items
                  if y_bottom <= t['y'] < y_top - 3
                  and t['layer'] == 'AM_4'
                  and t['text'] not in COLUMN_NAMES]  # 排除其他标题

    # 按 Y 聚类（行）
    sorted_data = sorted(data_texts, key=lambda t: -t['y'])
    rows = []
    if sorted_data:
        current = [sorted_data[0]]
        for t in sorted_data[1:]:
            if abs(t['y'] - current[-1]['y']) < 3.0:
                current.append(t)
            else:
                current.sort(key=lambda t: t['x'])
                rows.append(current)
                current = [t]
        current.sort(key=lambda t: t['x'])
        rows.append(current)

    tables.append({
        'y_top': y_top,
        'y_bottom': y_bottom,
        'col_x_map': col_x_map,
        'data_rows': rows,
        'header_count': len(grp),
    })

print(f"\n共 {len(tables)} 个图框表格")
for i, t in enumerate(tables):
    print(f"  图框 {i+1}: Y={t['y_top']:.0f}~{t['y_bottom']:.0f}, 数据行 {len(t['data_rows'])}")

# ===== 第四步：输出 Excel =====
wb = Workbook()
wb.remove(wb.active)

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="305496")
thin = Side(border_style="thin", color="999999")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
data_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center")

def assign_to_col(t_x, col_x_map, threshold=20):
    """数据行文字 → 列归属"""
    best_col = None
    best_dist = threshold
    for col, x in col_x_map.items():
        d = abs(t_x - x)
        if d < best_dist:
            best_dist = d
            best_col = col
    return best_col

# Sheet 1: 全部表格汇总
ws = wb.create_sheet("全部表格汇总")
ws.append(['图框#', 'Y范围', '标题行数', '行号'] + COLUMN_NAMES)
for c in ws[1]:
    c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = border

for t_idx, tbl in enumerate(tables):
    for r_idx, row in enumerate(tbl['data_rows']):
        row_data = {c: '' for c in COLUMN_NAMES}
        for t in row:
            col = assign_to_col(t['x'], tbl['col_x_map'])
            if col and t['text']:
                if row_data[col]:
                    row_data[col] += ' ' + t['text']
                else:
                    row_data[col] = t['text']
        ws.append([t_idx + 1, f"{tbl['y_top']:.0f}~{tbl['y_bottom']:.0f}",
                   tbl['header_count'], r_idx + 1] +
                  [row_data[c] for c in COLUMN_NAMES])
        r = ws.max_row
        for c_idx in range(1, 15):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = border; cell.alignment = data_align

for col, w in [('A', 8), ('B', 16), ('C', 10), ('D', 6),
               ('E', 25), ('F', 25), ('G', 15), ('H', 12), ('I', 8),
               ('J', 8), ('K', 8), ('L', 25), ('M', 15), ('N', 8)]:
    ws.column_dimensions[col].width = w
ws.freeze_panes = "E2"
ws.auto_filter.ref = ws.dimensions

# 每个图框单独 Sheet
for t_idx, tbl in enumerate(tables):
    sheet_name = f"图框{t_idx+1}"
    if len(sheet_name) > 31:
        sheet_name = f"T{t_idx+1}"
    ws = wb.create_sheet(sheet_name)
    ws.append(COLUMN_NAMES)
    for c in ws[1]:
        c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = border

    for row in tbl['data_rows']:
        row_data = {c: '' for c in COLUMN_NAMES}
        for t in row:
            col = assign_to_col(t['x'], tbl['col_x_map'])
            if col and t['text']:
                if row_data[col]:
                    row_data[col] += ' ' + t['text']
                else:
                    row_data[col] = t['text']
        ws.append([row_data[c] for c in COLUMN_NAMES])
        r = ws.max_row
        for c_idx in range(1, 11):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = border; cell.alignment = data_align

    for col, w in [('A', 25), ('B', 25), ('C', 15), ('D', 12), ('E', 8),
                  ('F', 8), ('G', 8), ('H', 25), ('I', 15), ('J', 8)]:
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

wb.save(OUTPUT_XLSX)
print(f"\n✅ Excel 已保存: {OUTPUT_XLSX}")
print(f"   大小: {OUTPUT_XLSX.stat().st_size/1024:.1f} KB")
print(f"\nSheet 列表:")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"  {name}: {ws.max_row-1} 行 x {ws.max_column} 列")
