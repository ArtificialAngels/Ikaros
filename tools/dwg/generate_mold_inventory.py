"""
江铃模具 2D 文件清单 - v2 (Bash 版本)
使用 Python + 直接遍历 + 之前验证过的解析规则
增强：扫描所有扩展名（不只 .dwg/.dwl/.dwl2/.bak），并把上一版漏掉的 6 个文件找回来
"""
import re
import os
from pathlib import Path
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r"E:\KPSNC模具资料\江铃\江铃缸体GE266 2.3 AN\ge266_2.3an(tac2)压铸模具2D,3D图档数据20250929\ge266_2.3an(tac2)压铸模具2D,3D图档数据20250929\2D")
OUTPUT = ROOT / "模具文件清单.xlsx"

# 项目号前缀（按长度从长到短排列，避免 L250864 误吃 25-8777）
PROJECT_PREFIXES = [
    r"L250864-C1411\s*\(25-8777\)",  # 模架图纸
    r"225-8777",                       # 偶发变体
    r"25-8777",
    r"25-8763",
]
PREFIX_PATTERN = re.compile(
    r"^(" + "|".join(PROJECT_PREFIXES) + r")[\s_\-]*",
    re.IGNORECASE
)
NUM_SUFFIX_PATTERN = re.compile(r"_(\d{4,5})$")
HAS_CHINESE = re.compile(r"[\u4e00-\u9fff]")
TEMP_EXTS = {"dwl", "dwl2", "bak"}
# 不计入清单的纯系统文件
SKIP_NAMES = {"Thumbs.db", ".DS_Store"}


def parse_filename(stem: str) -> dict:
    original = stem
    s = PREFIX_PATTERN.sub("", stem, count=1)
    prefix_removed = s

    # 1) 优先取末尾 4-5 位编号
    m = NUM_SUFFIX_PATTERN.search(s)
    number = ""
    desc_part = s
    if m:
        number = m.group(1)
        desc_part = s[:m.start()]
    else:
        # 2) 备选：找中间任意 _NNNN 段
        any_num = re.search(r"_(\d{4,5})\b", s)
        if any_num:
            number = any_num.group(1)
            desc_part = s[:any_num.start()]

    desc_part = re.sub(r"[\s_\-]+$", "", desc_part)

    # 取最后一个含中文的 _ 分段
    if HAS_CHINESE.search(desc_part):
        parts = desc_part.split("_")
        for i in range(len(parts) - 1, -1, -1):
            if HAS_CHINESE.search(parts[i]):
                description = parts[i]
                break
        else:
            description = desc_part
    else:
        description = desc_part

    description = description.strip()
    return {
        "original": original,
        "prefix_removed": prefix_removed,
        "description": description,
        "number": number,
    }


def main():
    print(f"扫描根目录: {ROOT}")
    if not ROOT.exists():
        raise SystemExit(f"❌ 目录不存在: {ROOT}")

    rows = []
    subdirs = sorted([p for p in ROOT.iterdir() if p.is_dir()])
    print(f"发现 {len(subdirs)} 个子目录")

    total_files = 0
    for sd in subdirs:
        files = sorted([p for p in sd.iterdir() if p.is_file()])
        print(f"  📁 {sd.name}: {len(files)} 个文件")
        for fp in files:
            # 跳过系统文件
            if fp.name in SKIP_NAMES:
                continue
            total_files += 1
            stem = fp.stem
            ext = fp.suffix.lower().lstrip(".")

            parsed = parse_filename(stem)

            remarks = []
            if ext in TEMP_EXTS:
                ext_name_map = {"dwl": "AutoCAD 锁文件", "dwl2": "AutoCAD 锁文件", "bak": "备份文件"}
                remarks.append(ext_name_map.get(ext, "临时文件"))
            if ext in ("rar", "zip", "7z"):
                remarks.append("压缩包")
            if ext in ("stp", "x_t", "step", "igs", "iges"):
                remarks.append("3D 中性格式")
            if not parsed["number"]:
                remarks.append("无编号")
            if not parsed["description"]:
                remarks.append("无中文描述")

            rows.append({
                "序号": total_files,
                "子目录": sd.name,
                "原文件名": fp.name,
                "中文描述": parsed["description"],
                "编号": parsed["number"],
                "扩展名": ext,
                "备注": "; ".join(remarks),
            })

    print(f"\n共解析 {total_files} 个文件")

    # ============ 写 Excel ============
    wb = Workbook()
    ws = wb.active
    ws.title = "模具文件清单"

    headers = ["序号", "子目录", "原文件名", "中文描述", "编号", "扩展名", "备注"]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="305496")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="999999")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    data_align = Alignment(horizontal="left", vertical="center", wrap_text=False)

    for col_idx, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border

    for r in rows:
        ws.append([r["序号"], r["子目录"], r["原文件名"], r["中文描述"], r["编号"], r["扩展名"], r["备注"]])
        row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = data_align
            cell.border = border
        # 临时文件行加灰底
        if r["扩展名"] in {"dwl", "dwl2", "bak"}:
            gray = PatternFill("solid", fgColor="F2F2F2")
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = gray
        # 3D 中间文件行加浅蓝
        if r["扩展名"] in {"stp", "x_t", "step", "igs", "iges"}:
            blue = PatternFill("solid", fgColor="DEEBF7")
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = blue
        # 压缩包行加浅黄
        if r["扩展名"] in {"rar", "zip", "7z"}:
            yellow = PatternFill("solid", fgColor="FFF2CC")
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = yellow

    col_widths = {1: 6, 2: 28, 3: 60, 4: 40, 5: 10, 6: 10, 7: 22}
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ============ 摘要 Sheet ============
    ws2 = wb.create_sheet("子目录统计")
    ws2.append(["子目录", "文件总数", "有编号", "无编号", "占总数%"])
    for col_idx in range(1, 6):
        c = ws2.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border

    sub_counter = Counter(r["子目录"] for r in rows)
    sub_with_num = Counter(r["子目录"] for r in rows if r["编号"])
    total_n = len(rows)
    for sub, cnt in sub_counter.most_common():
        with_n = sub_with_num[sub]
        no_n = cnt - with_n
        pct = cnt * 100 / total_n
        ws2.append([sub, cnt, with_n, no_n, f"{pct:.1f}%"])
        for col_idx in range(1, 6):
            cell = ws2.cell(row=ws2.max_row, column=col_idx)
            cell.border = border
            cell.alignment = data_align

    ws2.append([])
    ws2.append(["合计", total_n, sum(sub_with_num.values()), total_n - sum(sub_with_num.values()), "100.0%"])
    for col_idx in range(1, 6):
        cell = ws2.cell(row=ws2.max_row, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.border = border

    for col, w in [(1, 30), (2, 12), (3, 10), (4, 10), (5, 12)]:
        ws2.column_dimensions[get_column_letter(col)].width = w

    # ============ 扩展名统计 Sheet ============
    ws3 = wb.create_sheet("扩展名统计")
    ws3.append(["扩展名", "文件数", "占比%"])
    for col_idx in range(1, 4):
        c = ws3.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border

    ext_counter = Counter(r["扩展名"] for r in rows)
    for ext, cnt in ext_counter.most_common():
        pct = cnt * 100 / total_n
        ws3.append([ext, cnt, f"{pct:.1f}%"])
        for col_idx in range(1, 4):
            cell = ws3.cell(row=ws3.max_row, column=col_idx)
            cell.border = border
            cell.alignment = data_align
    for col, w in [(1, 15), (2, 10), (3, 10)]:
        ws3.column_dimensions[get_column_letter(col)].width = w

    wb.save(OUTPUT)
    print(f"\n✅ Excel 已保存: {OUTPUT}")
    print(f"   文件大小: {OUTPUT.stat().st_size / 1024:.1f} KB")

    # 抽检输出
    print("\n=== 抽检前 10 行 ===")
    print("-" * 110)
    for r in rows[:10]:
        print(f"  [{r['序号']:>3}] {r['子目录']:<25} | 描述: {r['中文描述']:<30} | 编号: {r['编号']:<6} | {r['原文件名']}")

    # 抽检结尾
    print("\n=== 抽检末尾 5 行 ===")
    print("-" * 110)
    for r in rows[-5:]:
        print(f"  [{r['序号']:>3}] {r['子目录']:<25} | 描述: {r['中文描述']:<30} | 编号: {r['编号']:<6} | {r['原文件名']}")


if __name__ == "__main__":
    main()
