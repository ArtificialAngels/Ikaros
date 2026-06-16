"""
完整提取多图框表格 - v4
策略：
1. 用"件号"或"冷却编号"的 X 位置作为图框的 X 锚点
2. 把相近 X 的标题行归为同一图框
3. 每个图框提取：标题行 + 向下找数据行
4. 用列 X 位置聚类数据
"""
import ezdxf
from pathlib import Path
from collections import Counter, defaultdict
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DXF_PATH = Path(r'E:\KPSNC模具资料\GEELY_BHE20_ICE\2026-04-08_25-8715_Geely BHE20-ICE\1.2D3DMOLd\2D\COOLING\25-8715_点冷却器（已打印）.dxf')
OUTPUT_XLSX = DXF_PATH.parent / "DXF表格提取.xlsx"

def clean_mtext(text: str) -> str:
    text = re.sub(r'\{[^}]*\}', '', text)
    text = re.sub(r'\\[WwFfTtCcPpQqAaLlKk][^;]*;', '', text)
    text = re.sub(r'\\[WwFfTtCcPpQqAaLlKk]', '', text)
    text = text.replace('%%c', 'Ø')
    text = re.sub(r'\s+', ' ', text).strip()
    return text

print(f"读取: {DXF_PATH.name}")
doc = ezdxf.readfile(str(DXF_PATH))
msp = doc.modelspace()
entities = list(msp)

# 收集所有文字
text_items = []
for e in entities:
    if e.dxftype() in ('TEXT', 'MTEXT'):
        try:
            raw = e.dxf.text if e.dxftype() == 'TEXT' else e.text
            if not raw or not raw.strip():
                continue
            clean = clean_mtext(raw)
            if not clean:
                continue
            pos = e.dxf.insert
            text_items.append({
                'x': float(pos[0]),
                'y': float(pos[1]),
                'text': clean,
                'raw': raw,
                'layer': e.dxf.layer,
            })
        except:
            pass

print(f"文字总数: {len(text_items)}")

# ===== 找所有"件号"作为图框锚点（必须是 AM_4 图层）=====
anchors = []
for t in text_items:
    if t['text'] == '件号' and t['layer'] == 'AM_4':
        anchors.append({'x': t['x'], 'y': t['y']})

print(f"\n图框锚点: {len(anchors)} 个")
for a in anchors:
    print(f"  ({a['x']:.1f}, {a['y']:.1f})")

# ===== 找所有标题行（同一图框可能有多套标题 - 大图/小图） =====
# 同一图框的标题行：X 相同，Y 接近
# 把相近 (X, Y) 的件号视为同一图框
def cluster_points(points, x_tol=200, y_tol=20):
    """聚类 2D 点"""
    if not points:
        return []
    clusters = [[points[0]]]
    for p in points[1:]:
        found = False
        for c in clusters:
            ref = c[0]
            if abs(p['x'] - ref['x']) < x_tol and abs(p['y'] - ref['y']) < y_tol:
                c.append(p)
                found = True
                break
        if not found:
            clusters.append([p])
    # 每簇取平均位置
    result = []
    for c in clusters:
        avg_x = sum(p['x'] for p in c) / len(c)
        avg_y = sum(p['y'] for p in c) / len(c)
        result.append({'x': avg_x, 'y': avg_y, 'count': len(c)})
    return result

clustered_anchors = cluster_points(anchors, x_tol=200, y_tol=20)
print(f"\n聚类后图框: {len(clustered_anchors)}")
for c in clustered_anchors:
    print(f"  锚点 ({c['x']:.1f}, {c['y']:.1f})  含 {c['count']} 个标题")

# ===== 对每个图框锚点，提取完整表格 =====
COLUMN_NAMES = ['件号', '冷却编号', '型号', 'd', 'L', 'E', 'J', '使用零件', '运水孔K', '数量']

def find_table_at_anchor(anchor_x, anchor_y, text_items):
    """以锚点为中心找表格"""
    # X 范围：锚点 X ± 500mm
    x_left = anchor_x - 30
    x_right = anchor_x + 600

    # Y 范围：向下到下一图框或 350mm 距离（典型表格高度）
    # 但实际上无法预测下一图框，先取一个保守的范围
    y_top = anchor_y + 5
    y_bottom = anchor_y - 350  # 一个图框典型高度

    # 找这个范围内的所有文字
    in_range = [t for t in text_items
                if y_bottom <= t['y'] <= y_top and x_left <= t['x'] <= x_right
                and t['layer'] in ('AM_4', '12')]

    # 找标题行（Y 最高的文字组）
    sorted_y = sorted(in_range, key=lambda t: -t['y'])
    if not sorted_y:
        return None

    # 标题行：Y 与最高 Y 相差 1mm 内
    top_y = sorted_y[0]['y']
    header = [t for t in sorted_y if abs(t['y'] - top_y) < 1.5]
    header.sort(key=lambda t: t['x'])

    # 列 X 位置（按 COLUMN_NAMES 顺序分配）
    col_x_map = {}
    for i, t in enumerate(header):
        if i < len(COLUMN_NAMES):
            col_x_map[COLUMN_NAMES[i]] = t['x']

    # 数据行：Y < 标题 Y
    data_texts = [t for t in sorted_y if t['y'] < top_y - 3]
    # 按 Y 聚类
    rows = []
    if data_texts:
        current = [data_texts[0]]
        for t in data_texts[1:]:
            if abs(t['y'] - current[-1]['y']) < 3.0:
                current.append(t)
            else:
                current.sort(key=lambda t: t['x'])
                rows.append(current)
                current = [t]
        current.sort(key=lambda t: t['x'])
        rows.append(current)

    return {
        'anchor_x': anchor_x,
        'anchor_y': anchor_y,
        'header': header,
        'col_x_map': col_x_map,
        'data_rows': rows,
    }

# 处理每个图框
tables = []
for a in clustered_anchors:
    tbl = find_table_at_anchor(a['x'], a['y'], text_items)
    if tbl:
        tables.append(tbl)

print(f"\n成功识别 {len(tables)} 个表格")
for i, t in enumerate(tables):
    print(f"  表格 {i+1}: 锚点 ({t['anchor_x']:.0f}, {t['anchor_y']:.0f}), 数据 {len(t['data_rows'])} 行")

# ===== 输出 Excel =====
wb = Workbook()
wb.remove(wb.active)

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="305496")
thin = Side(border_style="thin", color="999999")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
data_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center")

def assign_to_col(t_x, col_x_map, threshold=40):
    best_col = None
    best_dist = threshold
    for col, x in col_x_map.items():
        d = abs(t_x - x)
        if d < best_dist:
            best_dist = d
            best_col = col
    return best_col

# Sheet 1: 全部表格汇总
ws = wb.create_sheet("全部表格汇总")
ws.append(['图框#', '锚点 X', '锚点 Y', '行号'] + COLUMN_NAMES)
for c in ws[1]:
    c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = border

for t_idx, tbl in enumerate(tables):
    for r_idx, row in enumerate(tbl['data_rows']):
        row_data = {c: '' for c in COLUMN_NAMES}
        for t in row:
            col = assign_to_col(t['x'], tbl['col_x_map'])
            if col and t['text']:
                if row_data[col]:
                    row_data[col] += ' ' + t['text']
                else:
                    row_data[col] = t['text']
        ws.append([t_idx + 1, round(tbl['anchor_x']), round(tbl['anchor_y']), r_idx + 1] +
                  [row_data[c] for c in COLUMN_NAMES])
        r = ws.max_row
        for c_idx in range(1, 15):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = border; cell.alignment = data_align

for col, w in [('A', 8), ('B', 10), ('C', 10), ('D', 6),
               ('E', 25), ('F', 25), ('G', 15), ('H', 12), ('I', 8),
               ('J', 8), ('K', 8), ('L', 25), ('M', 15), ('N', 8)]:
    ws.column_dimensions[col].width = w
ws.freeze_panes = "E2"
ws.auto_filter.ref = ws.dimensions

# ===== 每个图框单独 Sheet =====
for t_idx, tbl in enumerate(tables):
    sheet_name = f"图框{t_idx+1}"
    if len(sheet_name) > 31:
        sheet_name = f"T{t_idx+1}"
    ws = wb.create_sheet(sheet_name)
    ws.append(COLUMN_NAMES)
    for c in ws[1]:
        c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = border

    for row in tbl['data_rows']:
        row_data = {c: '' for c in COLUMN_NAMES}
        for t in row:
            col = assign_to_col(t['x'], tbl['col_x_map'])
            if col and t['text']:
                if row_data[col]:
                    row_data[col] += ' ' + t['text']
                else:
                    row_data[col] = t['text']
        ws.append([row_data[c] for c in COLUMN_NAMES])
        r = ws.max_row
        for c_idx in range(1, 11):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = border; cell.alignment = data_align

    for col, w in [('A', 25), ('B', 25), ('C', 15), ('D', 12), ('E', 8),
                  ('F', 8), ('G', 8), ('H', 25), ('I', 15), ('J', 8)]:
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

wb.save(OUTPUT_XLSX)
print(f"\n✅ Excel 已保存: {OUTPUT_XLSX}")
print(f"   大小: {OUTPUT_XLSX.stat().st_size/1024:.1f} KB")
print(f"\nSheet 列表:")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"  {name}: {ws.max_row-1} 行 x {ws.max_column} 列")
