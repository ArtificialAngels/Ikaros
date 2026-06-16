"""
完整提取多图框表格 - v3
策略：
1. 用标题行的每个文字 X 位置作为"列分隔线"
2. 数据行的每个文字，按 X 位置归属到最近的"列"
3. 标题行的字段名按 X 顺序就是 [件号, 冷却编号, 型号, d, L, E, J, 使用零件, 运水孔K, 数量]
"""
import ezdxf
from pathlib import Path
from collections import Counter, defaultdict
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DXF_PATH = Path(r'E:\KPSNC模具资料\GEELY_BHE20_ICE\2026-04-08_25-8715_Geely BHE20-ICE\1.2D3DMOLd\2D\COOLING\25-8715_点冷却器（已打印）.dxf')
OUTPUT_XLSX = DXF_PATH.parent / "DXF表格提取.xlsx"

def clean_mtext(text: str) -> str:
    text = re.sub(r'\{[^}]*\}', '', text)
    text = re.sub(r'\\[WwFfTtCcPpQqAaLlKk][^;]*;', '', text)
    text = re.sub(r'\\[WwFfTtCcPpQqAaLlKk]', '', text)
    # 处理 %%c -> Ø
    text = text.replace('%%c', 'Ø')
    text = re.sub(r'\s+', ' ', text).strip()
    return text

print(f"读取: {DXF_PATH.name}")
doc = ezdxf.readfile(str(DXF_PATH))
msp = doc.modelspace()
entities = list(msp)

# 收集所有文字
text_items = []
for e in entities:
    if e.dxftype() in ('TEXT', 'MTEXT'):
        try:
            raw = e.dxf.text if e.dxftype() == 'TEXT' else e.text
            if not raw or not raw.strip():
                continue
            clean = clean_mtext(raw)
            if not clean:
                continue
            pos = e.dxf.insert
            text_items.append({
                'x': float(pos[0]),
                'y': float(pos[1]),
                'text': clean,
                'raw': raw,
                'layer': e.dxf.layer,
            })
        except:
            pass

print(f"文字总数: {len(text_items)}")

# ===== 找标题行（Y 容差 1mm 内必须同时含"件号"和"冷却编号"）=====
HEADER_KEYS = {'件号', '冷却编号', '型号', '使用零件', '数量'}

# 按 Y 聚类
sorted_by_y = sorted(text_items, key=lambda t: -t['y'])
y_groups = []
current = [sorted_by_y[0]]
for t in sorted_by_y[1:]:
    if abs(t['y'] - current[-1]['y']) < 1.0:
        current.append(t)
    else:
        y_groups.append(current)
        current = [t]
y_groups.append(current)

# 找标题行（含全部 5 个核心关键词的 Y 簇）
header_y = None
for grp in y_groups:
    texts = [t['text'] for t in grp]
    if all(k in texts for k in ['件号', '冷却编号', '型号', '使用零件', '数量']):
        header_y = grp[0]['y']
        break

print(f"第一个标题行 Y={header_y}")

# 找所有标题行（与首个标题 Y 距离分组）
# 一个图框通常 Y 范围在 ~80-200mm 内
# 方法：找所有包含全部 5 个关键词的 Y 簇
header_locs = []
for grp in y_groups:
    texts = [t['text'] for t in grp]
    if all(k in texts for k in ['件号', '冷却编号', '型号', '使用零件', '数量']):
        # 还要是 AM_4 图层
        if all(t['layer'] == 'AM_4' for t in grp):
            header_locs.append(grp[0]['y'])

# 按 Y 排序
header_locs.sort(reverse=True)  # AutoCAD Y 大在上
print(f"\n共找到 {len(header_locs)} 个标题行")
print(f"Y 范围: {min(header_locs):.0f} ~ {max(header_locs):.0f}")

# 计算相邻标题的 Y 差，找"图框边界"
# 一个图框的 Y 范围通常 ~100-300mm
print("\n相邻标题 Y 差（前 20）:")
diffs = [header_locs[i] - header_locs[i+1] for i in range(len(header_locs)-1)]
for i, d in enumerate(diffs[:20]):
    print(f"  {header_locs[i]:.0f} -> {header_locs[i+1]:.0f}: 差 {d:.1f}")

# ===== 聚类标题行为"图框组" =====
# Y 差 < 30mm 视为同一图框的不同小段（有些图纸有多段）
# Y 差 >= 30mm 视为新图框
TABLE_GAP = 30  # mm

# 合并相近的标题
merged_header_groups = []  # [[y1, y2, ...], ...]
current_group = [header_locs[0]]
for y in header_locs[1:]:
    if current_group[-1] - y < TABLE_GAP:
        current_group.append(y)
    else:
        merged_header_groups.append(current_group)
        current_group = [y]
merged_header_groups.append(current_group)

print(f"\n合并后: {len(merged_header_groups)} 个图框")

# ===== 对每个图框组，提取所有文字 =====
# 每个图框的 Y 范围:
#   - 上: 最高标题 Y + 5
#   - 下: 下一图框的最低标题 Y - 30 (留空)
def get_table_y_range(group, next_group_start):
    y_top = max(group) + 5
    y_bottom = next_group_start - 30 if next_group_start else -100000
    return y_top, y_bottom

# 标准列名（按 X 顺序）
COLUMN_NAMES = ['件号', '冷却编号', '型号', 'd', 'L', 'E', 'J', '使用零件', '运水孔K', '数量']

tables = []
for i, group in enumerate(merged_header_groups):
    y_top, y_bottom = get_table_y_range(group, merged_header_groups[i+1][0] if i+1 < len(merged_header_groups) else None)

    # 在图框 Y 范围内找所有文字
    in_table = [t for t in text_items if y_bottom <= t['y'] <= y_top and t['layer'] in ('AM_4', '12', '标注', 'AM_5', '0', 'TEXT', 'DIM')]

    # 找标题行（Y 最高的）
    sorted_in = sorted(in_table, key=lambda t: -t['y'])
    header_row = [t for t in sorted_in if t['y'] > y_top - 10]
    # 取 X 排序
    header_row.sort(key=lambda t: t['x'])

    # 找每个标题文字的 X（标准列顺序）
    # 用第一个标题（X 最小的"件号"）的位置作为基准
    col_x_map = {}  # col_name -> x
    # 按 X 顺序给每个 header 分配列名
    for j, t in enumerate(header_row):
        if j < len(COLUMN_NAMES):
            col_x_map[COLUMN_NAMES[j]] = t['x']
    if not col_x_map:
        continue

    # 找数据行：Y < 标题 Y（更靠下）
    data_rows_y = [t for t in sorted_in if t['y'] < y_top - 10]
    # 按 Y 聚类
    data_rows_grouped = []
    if data_rows_y:
        current = [data_rows_y[0]]
        for t in data_rows_y[1:]:
            if abs(t['y'] - current[-1]['y']) < 3.0:
                current.append(t)
            else:
                current.sort(key=lambda t: t['x'])
                data_rows_grouped.append(current)
                current = [t]
        current.sort(key=lambda t: t['x'])
        data_rows_grouped.append(current)

    # 收集所有行（包括数据行）
    all_rows = []
    # 标题行作为第一行
    all_rows.append([{'col': COLUMN_NAMES[i] if i < len(COLUMN_NAMES) else f'col{i}',
                      'text': t['text']} for i, t in enumerate(header_row)])

    for row in data_rows_grouped:
        all_rows.append([{'col': '?', 'text': t['text'], 'x': t['x']} for t in row])

    tables.append({
        'y_top': y_top,
        'y_bottom': y_bottom,
        'header_row': header_row,
        'col_x_map': col_x_map,
        'data_rows': data_rows_grouped,
    })

print(f"\n最终识别 {len(tables)} 个图框表格")
for i, t in enumerate(tables):
    print(f"  图框 {i+1}: Y={t['y_top']:.0f} ~ {t['y_bottom']:.0f}, 数据行 {len(t['data_rows'])}")

# ===== 输出 Excel =====
wb = Workbook()
wb.remove(wb.active)

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="305496")
thin = Side(border_style="thin", color="999999")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
data_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center")

def assign_to_col(t_x, col_x_map, threshold=40):
    """根据 X 位置归属到最近的列"""
    best_col = None
    best_dist = threshold
    for col, x in col_x_map.items():
        d = abs(t_x - x)
        if d < best_dist:
            best_dist = d
            best_col = col
    return best_col

# Sheet 1: 全部表格汇总
ws = wb.create_sheet("全部表格汇总")
ws.append(['图框#', 'Y范围', '行号'] + COLUMN_NAMES)
for c in ws[1]:
    c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = border

for t_idx, tbl in enumerate(tables):
    for r_idx, row in enumerate(tbl['data_rows']):
        row_data = {c: '' for c in COLUMN_NAMES}
        for t in row:
            col = assign_to_col(t['x'], tbl['col_x_map'])
            if col and t['text']:
                if row_data[col]:
                    row_data[col] += ' ' + t['text']
                else:
                    row_data[col] = t['text']
        ws.append([t_idx + 1, f"{tbl['y_top']:.0f}~{tbl['y_bottom']:.0f}", r_idx + 1] +
                  [row_data[c] for c in COLUMN_NAMES])
        r = ws.max_row
        for c_idx in range(1, 14):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = border; cell.alignment = data_align

for col, w in [('A', 8), ('B', 14), ('C', 6), ('D', 25), ('E', 25), ('F', 15),
               ('G', 12), ('H', 8), ('I', 8), ('J', 8), ('K', 8),
               ('L', 25), ('M', 15), ('N', 8)]:
    ws.column_dimensions[col].width = w
ws.freeze_panes = "D2"
ws.auto_filter.ref = ws.dimensions

# ===== 每个图框单独 Sheet =====
for t_idx, tbl in enumerate(tables):
    sheet_name = f"图框{t_idx+1}"
    if len(sheet_name) > 31:
        sheet_name = f"T{t_idx+1}"
    ws = wb.create_sheet(sheet_name)
    ws.append(COLUMN_NAMES)
    for c in ws[1]:
        c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = border

    for row in tbl['data_rows']:
        row_data = {c: '' for c in COLUMN_NAMES}
        for t in row:
            col = assign_to_col(t['x'], tbl['col_x_map'])
            if col and t['text']:
                if row_data[col]:
                    row_data[col] += ' ' + t['text']
                else:
                    row_data[col] = t['text']
        ws.append([row_data[c] for c in COLUMN_NAMES])
        r = ws.max_row
        for c_idx in range(1, 11):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = border; cell.alignment = data_align

    for col, w in [('A', 25), ('B', 25), ('C', 15), ('D', 12), ('E', 8),
                  ('F', 8), ('G', 8), ('H', 25), ('I', 15), ('J', 8)]:
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

wb.save(OUTPUT_XLSX)
print(f"\n✅ Excel 已保存: {OUTPUT_XLSX}")
print(f"   大小: {OUTPUT_XLSX.stat().st_size/1024:.1f} KB")
print(f"\nSheet 列表:")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"  {name}: {ws.max_row-1} 行 x {ws.max_column} 列")
