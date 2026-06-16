"""
第三步：质量抽检 - 验证 Excel 解析结果
"""
import re
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(r"E:\KPSNC模具资料\江铃\江铃缸体GE266 2.3 AN\ge266_2.3an(tac2)压铸模具2D,3D图档数据20250929\ge266_2.3an(tac2)压铸模具2D,3D图档数据20250929\2D")
XLSX = ROOT / "模具文件清单.xlsx"

wb = load_workbook(XLSX)
ws = wb.active

print(f"工作表: {ws.title}")
print(f"总行数: {ws.max_row}")
print(f"总列数: {ws.max_column}")
print()

# 解析每行
rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    # 把 None 转为空字符串
    r = tuple("" if v is None else v for v in r)
    rows.append(r)

# 统计
total = len(rows)
with_number = sum(1 for r in rows if r[4])  # 编号
with_desc = sum(1 for r in rows if r[3])    # 中文描述
no_number = sum(1 for r in rows if not r[4])
no_desc = sum(1 for r in rows if not r[3])

print(f"总文件数: {total}")
print(f"有编号的: {with_number} ({with_number*100//total}%)")
print(f"有中文描述的: {with_desc} ({with_desc*100//total}%)")
print(f"无编号的: {no_number} ({no_number*100//total}%)")
print(f"无中文描述的: {no_desc} ({no_desc*100//total}%)")
print()

# 按子目录统计
print("=== 按子目录统计 ===")
from collections import Counter
subdir_counter = Counter(r[1] for r in rows)
for sub, cnt in subdir_counter.most_common():
    sub_rows = [r for r in rows if r[1] == sub]
    sub_with_num = sum(1 for r in sub_rows if r[4])
    print(f"  {sub:<30} : {cnt:>3} 个文件, 其中 {sub_with_num:>3} 个有编号")

print()
print("=== 抽检 15 条典型记录（混合）===")
# 取每个子目录的前 1-2 条
shown = set()
for r in rows:
    if r[1] not in shown:
        shown.add(r[1])
        print(f"  [{r[0]:>3}] {r[1]:<25} | 描述: {r[3]:<30} | 编号: {r[4]:<6} | 扩展名: {r[5]:<5} | 文件: {r[2]}")
    if len(shown) >= 17:
        break

print()
print("=== 抽检 5 条'无编号'记录 ===")
no_num_rows = [r for r in rows if not r[4]]
for r in no_num_rows[:5]:
    print(f"  [{r[0]:>3}] {r[1]:<25} | 描述: {r[3]:<30} | 文件: {r[2]}")

print()
print("=== 抽检 5 条'无中文描述'记录 ===")
no_desc_rows = [r for r in rows if not r[3]]
for r in no_desc_rows[:5]:
    print(f"  [{r[0]:>3}] {r[1]:<25} | 文件: {r[2]}")

# 检查重复的(描述,编号) 组合
print()
print("=== 检查可能的(描述,编号)重复 ===")
combos = Counter((r[3], r[4]) for r in rows if r[3] and r[4])
dups = [(c, n) for c, n in combos.items() if n > 1]
print(f"  共有 {len(dups)} 个 (描述,编号) 出现多次")
for c, n in dups[:5]:
    print(f"    {c} -> {n} 次")
