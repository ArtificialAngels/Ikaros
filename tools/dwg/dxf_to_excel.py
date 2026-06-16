"""
将 DXF 提取结果转为 Excel
包含：
- Sheet 1: 总体统计
- Sheet 2: 标题栏/签字栏/BOM (TabularNote 块)
- Sheet 3: 各图层文字（疑似表格内容）
- Sheet 4: 全部 MTEXT 文字（含格式）
- Sheet 5: 块插入位置
"""
import ezdxf
from pathlib import Path
from collections import Counter, defaultdict
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ===== 配置 =====
DXF_PATH = Path(r'E:\KPSNC模具资料\GEELY_BHE20_ICE\2026-04-08_25-8715_Geely BHE20-ICE\1.2D3DMOLd\2D\COOLING\25-8715_点冷却器（已打印）.dxf')
OUTPUT_XLSX = DXF_PATH.parent / "DXF内容提取.xlsx"

# ===== 工具 =====
def clean_mtext(text: str) -> str:
    """清理 MTEXT 格式控制符"""
    text = re.sub(r'\{[^}]*\}', '', text)
    text = re.sub(r'\\[WwFfTtCcPpQqAaLlKk][^;]*;', '', text)
    text = re.sub(r'\\[WwFfTtCcPpQqAaLlKk]', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

# ===== 读 DXF =====
print(f"读取: {DXF_PATH.name} ({DXF_PATH.stat().st_size/1024/1024:.2f} MB)")
doc = ezdxf.readfile(str(DXF_PATH))
msp = doc.modelspace()
entities = list(msp)

# ===== 准备样式 =====
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="305496")
header_align = Alignment(horizontal="center", vertical="center")
thin = Side(border_style="thin", color="999999")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
data_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()
wb.remove(wb.active)  # 删默认 sheet

# ========================================
# Sheet 1: 文件统计
# ========================================
ws = wb.create_sheet("文件统计")
ws.append(["项目", "值"])
for c in ws[1]:
    c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = border

stats = {
    "文件名": DXF_PATH.name,
    "文件大小(MB)": round(DXF_PATH.stat().st_size / 1024 / 1024, 2),
    "DXF 版本": doc.dxfversion,
    "ACAD 版本": doc.acad_release,
    "总实体数": len(entities),
    "图层数": len(list(doc.layers)),
    "块定义数": len(list(doc.blocks)),
    "INSERT 块插入数": len([e for e in entities if e.dxftype() == 'INSERT']),
    "MTEXT 多行文字": len([e for e in entities if e.dxftype() == 'MTEXT']),
    "TEXT 单行文字": len([e for e in entities if e.dxftype() == 'TEXT']),
    "LINE 线段": len([e for e in entities if e.dxftype() == 'LINE']),
    "LWPOLYLINE 多段线": len([e for e in entities if e.dxftype() == 'LWPOLYLINE']),
    "CIRCLE 圆": len([e for e in entities if e.dxftype() == 'CIRCLE']),
    "ARC 圆弧": len([e for e in entities if e.dxftype() == 'ARC']),
    "SPLINE 样条": len([e for e in entities if e.dxftype() == 'SPLINE']),
    "HATCH 填充": len([e for e in entities if e.dxftype() == 'HATCH']),
    "DIMENSION 尺寸标注": len([e for e in entities if e.dxftype() == 'DIMENSION']),
    "LEADER 引线": len([e for e in entities if e.dxftype() == 'LEADER']),
    "OLE2FRAME 嵌入对象": len([e for e in entities if e.dxftype() == 'OLE2FRAME']),
    "ACAD_TABLE 表格实体": len([e for e in entities if e.dxftype() in ('ACAD_TABLE', 'TABLE')]),
}
for k, v in stats.items():
    ws.append([k, v])
    row_idx = ws.max_row
    for col_idx in range(1, 3):
        c = ws.cell(row=row_idx, column=col_idx)
        c.border = border; c.alignment = data_align

ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 50
ws.freeze_panes = "A2"

# ========================================
# Sheet 2: 标题栏/签字栏/BOM (TabularNote 块)
# ========================================
ws = wb.create_sheet("标题栏_BOM")
ws.append(["块名", "字段", "内容", "类型"])
for c in ws[1]:
    c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = border

block_titles = {
    'TabularNote2': '标题栏',
    'TabularNote3': '签字栏',
    'TabularNote4': 'BOM表'
}
for block_name, title in block_titles.items():
    if block_name in doc.blocks:
        block = doc.blocks[block_name]
        for e in block:
            if e.dxftype() in ('TEXT', 'MTEXT'):
                try:
                    raw = e.dxf.text if e.dxftype() == 'TEXT' else e.text
                    clean = clean_mtext(raw)
                    if clean:
                        ws.append([block_name, title, clean, e.dxftype()])
                        r = ws.max_row
                        for c_idx in range(1, 5):
                            cell = ws.cell(row=r, column=c_idx)
                            cell.border = border; cell.alignment = data_align
                        # 浅色区分
                        if block_name == 'TabularNote2':
                            fill_color = "E7F3FE"
                        elif block_name == 'TabularNote3':
                            fill_color = "FFF4E6"
                        else:
                            fill_color = "E8F5E9"
                        for c_idx in range(1, 5):
                            ws.cell(row=r, column=c_idx).fill = PatternFill("solid", fgColor=fill_color)
                except:
                    pass

ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 80
ws.column_dimensions['D'].width = 10
ws.freeze_panes = "A2"

# ========================================
# Sheet 3: 各图层文字（疑似表格）
# ========================================
ws = wb.create_sheet("图层文字统计")
ws.append(["图层", "总文字数", "唯一文字数", "含中文", "示例文字"])
for c in ws[1]:
    c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = border

text_by_layer = defaultdict(list)
chinese_pattern = re.compile(r'[\u4e00-\u9fff]')
for e in entities:
    if e.dxftype() in ('TEXT', 'MTEXT'):
        try:
            raw = e.dxf.text if e.dxftype() == 'TEXT' else e.text
            clean = clean_mtext(raw)
            if clean:
                text_by_layer[e.dxf.layer].append(clean)
        except:
            pass

# 按总文字数排序
sorted_layers = sorted(text_by_layer.items(), key=lambda x: -len(x[1]))
for layer, texts in sorted_layers:
    unique = list(dict.fromkeys(texts))
    cn_count = sum(1 for t in unique if chinese_pattern.search(t))
    sample = " | ".join(unique[:5])
    if len(sample) > 200:
        sample = sample[:200] + "..."
    ws.append([layer, len(texts), len(unique), cn_count, sample])
    r = ws.max_row
    for c_idx in range(1, 6):
        cell = ws.cell(row=r, column=c_idx)
        cell.border = border; cell.alignment = data_align

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 80
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

# ========================================
# Sheet 4: 疑似表格内容（图层聚合）
# ========================================
ws = wb.create_sheet("疑似表格内容")
ws.append(["图层", "文字(清洗后)", "原文字(原始)", "文字类型", "是否含中文"])
for c in ws[1]:
    c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = border

# 关键图层（疑似表格的）
key_layers = ['AM_4', '12', '标注', 'AM_5', 'DIM', 'CAXA细实线层', 'TEXT', '图框层', '排图层', 'b表格']
for layer in key_layers:
    if layer not in text_by_layer:
        continue
    texts = text_by_layer[layer]
    for raw, clean in [(t, t) for t in texts]:
        has_cn = "✓" if chinese_pattern.search(clean) else "✗"
        ws.append([layer, clean, clean, "TEXT/MTEXT", has_cn])
        r = ws.max_row
        for c_idx in range(1, 6):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = border; cell.alignment = data_align

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 60
ws.column_dimensions['C'].width = 60
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 12
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

# ========================================
# Sheet 5: 全部 MTEXT 文字
# ========================================
ws = wb.create_sheet("全部MTEXT文字")
ws.append(["序号", "图层", "清洗后文字", "原始文字(前200字)"])
for c in ws[1]:
    c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = border

idx = 0
for e in entities:
    if e.dxftype() == 'MTEXT':
        try:
            raw = e.text
            if not raw or not raw.strip():
                continue
            clean = clean_mtext(raw)
            if not clean:
                continue
            idx += 1
            ws.append([idx, e.dxf.layer, clean, raw[:200]])
            r = ws.max_row
            for c_idx in range(1, 5):
                cell = ws.cell(row=r, column=c_idx)
                cell.border = border; cell.alignment = data_align
        except:
            pass

ws.column_dimensions['A'].width = 8
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 80
ws.column_dimensions['D'].width = 80
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

# ========================================
# Sheet 6: 块插入位置
# ========================================
ws = wb.create_sheet("块插入位置")
ws.append(["块名", "插入点 X", "插入点 Y", "插入点 Z", "图层", "块实体数"])
for c in ws[1]:
    c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = border

inserts = [e for e in entities if e.dxftype() == 'INSERT']
for e in inserts:
    loc = e.dxf.insert
    block_name = e.dxf.name
    try:
        blk = doc.blocks.get(block_name)
        blk_ent_count = len(list(blk)) if blk else 0
    except:
        blk_ent_count = 0
    ws.append([block_name, round(loc[0], 2), round(loc[1], 2), round(loc[2], 2), e.dxf.layer, blk_ent_count])
    r = ws.max_row
    for c_idx in range(1, 7):
        cell = ws.cell(row=r, column=c_idx)
        cell.border = border; cell.alignment = data_align

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 12
ws.freeze_panes = "A2"

# ========================================
# Sheet 7: 全部图层清单
# ========================================
ws = wb.create_sheet("图层清单")
ws.append(["图层名", "颜色", "开/关", "线型"])
for c in ws[1]:
    c.font = header_font; c.fill = header_fill; c.alignment = header_align; c.border = border

for lyr in doc.layers:
    try:
        is_on = "开" if lyr.is_on() else "关"
    except:
        is_on = "?"
    try:
        ltype = lyr.dxf.linetype
    except:
        ltype = "?"
    ws.append([lyr.dxf.name, lyr.dxf.color, is_on, ltype])
    r = ws.max_row
    for c_idx in range(1, 5):
        cell = ws.cell(row=r, column=c_idx)
        cell.border = border; cell.alignment = data_align

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 15
ws.freeze_panes = "A2"

# ===== 保存 =====
wb.save(OUTPUT_XLSX)
print(f"\n✅ Excel 已保存: {OUTPUT_XLSX}")
print(f"   大小: {OUTPUT_XLSX.stat().st_size/1024:.1f} KB")
print(f"\n   Sheet 列表:")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"     {name}: {ws.max_row-1} 行 x {ws.max_column} 列")
