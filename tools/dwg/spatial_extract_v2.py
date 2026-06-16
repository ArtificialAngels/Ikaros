"""
完整提取多图框表格
策略：
1. 找所有"标题行"位置（Y 坐标）
2. 每个标题行向下扩展直到碰到下一个标题行（多个图框）
3. 用 X 坐标聚类列
4. 输出每个图框一张表
"""
import ezdxf
from pathlib import Path
from collections import Counter, defaultdict
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DXF_PATH = Path(r'E:\KPSNC模具资料\GEELY_BHE20_ICE\2026-04-08_25-8715_Geely BHE20-ICE\1.2D3DMOLd\2D\COOLING\25-8715_点冷却器（已打印）.dxf')
OUTPUT_XLSX = DXF_PATH.parent / "DXF表格提取.xlsx"

def clean_mtext(text: str) -> str:
    text = re.sub(r'\{[^}]*\}', '', text)
    text = re.sub(r'\\[WwFfTtCcPpQqAaLlKk][^;]*;', '', text)
    text = re.sub(r'\\[WwFfTtCcPpQqAaLlKk]', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

# 字段类型识别
DIMENSION_PATTERN = re.compile(r'^\d+(\.\d+)?$')  # 纯数字
DIAMETER_PATTERN = re.compile(r'^%%c\d|^\W?\d+\W*$|^\W?\d+(\.\d+)?\W*\W*\d+(\.\d+)?\W*$')  # %%c 表示直径
HYPHEN_NUMBER = re.compile(r'^\d+(~\d+)?$|^\d+-\d+(~\d+-\d+)?$|^\d+[~-]\d+[~-]\d+$')

print(f"读取: {DXF_PATH.name}")
doc = ezdxf.readfile(str(DXF_PATH))
msp = doc.modelspace()
entities = list(msp)

# 收集所有文字
text_items = []
for e in entities:
    if e.dxftype() in ('TEXT', 'MTEXT'):
        try:
            raw = e.dxf.text if e.dxftype() == 'TEXT' else e.text
            if not raw or not raw.strip():
                continue
            clean = clean_mtext(raw)
            if not clean:
                continue
            pos = e.dxf.insert
            try:
                h = float(e.dxf.get('height', 2.5))
            except:
                h = 2.5
            text_items.append({
                'x': float(pos[0]),
                'y': float(pos[1]),
                'h': h,
                'text': clean,
                'raw': raw,
                'layer': e.dxf.layer,
            })
        except:
            pass

print(f"文字总数: {len(text_items)}")

# ===== 找标题行（含'件号'或'冷却编号'且 Y 接近 0.5mm 容差）=====
header_y_set = set()
header_locs = []  # [(y, x_min, x_max)]
for t in text_items:
    if t['text'] in ('件号', '冷却编号') and t['layer'] in ('AM_4', '12'):
        # 找一行（容差 1mm）
        y_key = round(t['y'], 0)
        if y_key not in header_y_set:
            header_y_set.add(y_key)
            header_locs.append({'y': t['y'], 'x': t['x'], 'text': t['text']})

# 按 Y 排序
header_locs.sort(key=lambda h: -h['y'])
print(f"\n发现 {len(header_locs)} 个标题行位置")
for h in header_locs:
    print(f"  Y={h['y']:.1f}  X={h['x']:.1f}  起始文字: {h['text']}")

# ===== 收集 LINE 端点（用于确定每个表格的 X 范围）=====
line_endpoints = []
for e in entities:
    if e.dxftype() == 'LINE':
        try:
            s = e.dxf.start
            end = e.dxf.end
            line_endpoints.append({
                'x1': float(s[0]), 'y1': float(s[1]),
                'x2': float(end[0]), 'y2': float(end[1]),
                'layer': e.dxf.layer
            })
        except:
            pass

horizontal_lines = [l for l in line_endpoints if abs(l['y1'] - l['y2']) < 0.5 and abs(l['x2']-l['x1']) > 20]
print(f"\n长水平线: {len(horizontal_lines)}")

# ===== 找标题行的"列分割线" =====
# 对每个标题行，向上/向下 5mm 范围找垂直线 (dx ≈ 0)
def find_columns_at_y(y_target, lines, y_tol=3.0):
    """在 y_target 附近找垂直线（X 位置）"""
    x_positions = []
    for l in lines:
        if abs(l['x1'] - l['x2']) < 0.5:  # 垂直线
            mid_y = (l['y1'] + l['y2']) / 2
            if abs(mid_y - y_target) < y_tol:
                x = (l['x1'] + l['x2']) / 2
                # X 范围（线起止）
                x_min = min(l['y1'], l['y2'])
                x_max = max(l['y1'], l['y2'])
                x_positions.append({'x': x, 'y_min': min(l['y1'], l['y2']), 'y_max': max(l['y1'], l['y2'])})
    # 按 X 排序
    x_positions.sort(key=lambda p: p['x'])
    return x_positions

# 找所有垂直线
vertical_lines = [l for l in line_endpoints if abs(l['x1'] - l['x2']) < 0.5]
print(f"垂直线总数: {len(vertical_lines)}")

# ===== 对每个标题行，提取它的"表"区域 =====
# 关键列定义（按 X 位置从左到右）
COLUMN_NAMES = ['件号', '冷却编号', '型号', 'd', 'L', 'E', 'J', '使用零件', '运水孔K', '数量']

def cluster_x_positions(x_list, tol=2.0):
    """聚类 X 坐标"""
    if not x_list:
        return []
    x_sorted = sorted(x_list)
    clusters = [[x_sorted[0]]]
    for x in x_sorted[1:]:
        if abs(x - clusters[-1][-1]) < tol:
            clusters[-1].append(x)
        else:
            clusters.append([x])
    return [sum(c)/len(c) for c in clusters]

# 处理每个标题行
tables = []  # 每个元素是一个图框的表格
for i, h in enumerate(header_locs):
    y_header = h['y']
    x_start = h['x']

    # 找该标题行的"下一行"作为底部（也可能是另一个标题行）
    # 标题行之间的 Y 距离
    if i + 1 < len(header_locs):
        y_next = header_locs[i+1]['y']
    else:
        y_next = -10000  # 远下方

    # 在 y_header ~ y_next 之间找文字
    # 区域: y ∈ [y_next - 60, y_header + 5]   (AutoCAD Y 朝上, 大值在上)
    y_top = y_header + 5
    y_bottom = y_next - 50  # 距下一标题 50mm 以上

    # 在 X 范围 [x_start - 30, x_start + 1500] 找文字
    x_left = x_start - 50
    x_right = x_start + 1800

    in_table = [t for t in text_items
                if y_bottom <= t['y'] <= y_top and x_left <= t['x'] <= x_right
                and t['layer'] in ('AM_4', '12', '标注', 'AM_5', '0', 'TEXT', 'DIM')]

    # 按 Y 聚类
    sorted_by_y = sorted(in_table, key=lambda t: -t['y'])
    rows = []
    current = [sorted_by_y[0]]
    for t in sorted_by_y[1:]:
        if abs(t['y'] - current[-1]['y']) < 4.0:
            current.append(t)
        else:
            current.sort(key=lambda t: t['x'])
            rows.append(current)
            current = [t]
    current.sort(key=lambda t: t['x'])
    rows.append(current)

    # 第一行是标题行
    if not rows:
        continue

    header_row = rows[0]
    data_rows = rows[1:]

    # 收集列的 X 位置（标题行的 X）
    col_xs = [t['x'] for t in header_row]
    # 用列名做映射
    col_map = {}  # x -> col_name
    for t in header_row:
        col_map[t['x']] = t['text']

    # 收集所有数据
    table_data = {
        'header': header_row,
        'rows': data_rows,
        'col_xs': sorted(col_xs),
        'col_map': col_map,
        'y_header': y_header,
        'x_start': x_start,
    }
    tables.append(table_data)

print(f"\n共识别 {len(tables)} 个图框表格")
for i, tbl in enumerate(tables):
    print(f"  表格 {i+1}: 标题行 Y={tbl['y_header']:.1f}, 数据行 {len(tbl['rows'])} 行")

# ===== 输出 Excel =====
wb = Workbook()
wb.remove(wb.active)

# 样式
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="305496")
thin = Side(border_style="thin", color="999999")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
data_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
center_align = Alignment(horizontal="center", vertical="center")
yellow_fill = PatternFill("solid", fgColor="FFF2CC")

# Sheet 1: 全部表格汇总
ws = wb.create_sheet("全部表格汇总")
ws.append(["图框#", "表头行 Y", "起始X", "行号", "件号", "冷却编号", "型号", "d", "L", "E", "J", "使用零件", "运水孔K", "数量"])
for c in ws[1]:
    c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = border

current_table_idx = 0
for tbl in tables:
    current_table_idx += 1
    y_h = tbl['y_header']
    x_s = tbl['x_start']
    # 标题行
    header_row = tbl['header']
    header_texts = {t['x']: t['text'] for t in header_row}
    # 数据行
    for r_idx, row in enumerate(tbl['rows']):
        row_data = {'件号': '', '冷却编号': '', '型号': '', 'd': '', 'L': '', 'E': '', 'J': '',
                    '使用零件': '', '运水孔K': '', '数量': ''}
        for t in row:
            # 找最近的列
            best_col = None
            best_dist = 1e9
            for col in COLUMN_NAMES:
                if col not in header_texts.values():
                    continue
                # 找这个列的 X
                for hx, htext in header_texts.items():
                    if htext == col:
                        dist = abs(t['x'] - hx)
                        if dist < best_dist and dist < 50:
                            best_dist = dist
                            best_col = col
                        break
            if best_col and t['text']:
                if row_data[best_col]:
                    row_data[best_col] += ' ' + t['text']
                else:
                    row_data[best_col] = t['text']

        ws.append([current_table_idx, round(y_h, 1), round(x_s, 1), r_idx + 1,
                   row_data['件号'], row_data['冷却编号'], row_data['型号'],
                   row_data['d'], row_data['L'], row_data['E'], row_data['J'],
                   row_data['使用零件'], row_data['运水孔K'], row_data['数量']])
        r = ws.max_row
        for c_idx in range(1, 15):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = border; cell.alignment = data_align

# 标题行单独
for ti, tbl in enumerate(tables):
    pass  # 已经在数据中

for col_letter, w in [('A', 8), ('B', 12), ('C', 12), ('D', 6), ('E', 25), ('F', 25),
                       ('G', 15), ('H', 12), ('I', 8), ('J', 8), ('K', 8),
                       ('L', 25), ('M', 15), ('N', 8)]:
    ws.column_dimensions[col_letter].width = w
ws.freeze_panes = "E2"
ws.auto_filter.ref = ws.dimensions

# ===== 每个图框单独一个 Sheet =====
for tbl_idx, tbl in enumerate(tables):
    sheet_name = f"图框{tbl_idx+1}"
    if len(sheet_name) > 31:
        sheet_name = f"T{tbl_idx+1}"
    ws = wb.create_sheet(sheet_name)
    ws.append(['件号', '冷却编号', '型号', 'd', 'L', 'E', 'J', '使用零件', '运水孔K', '数量'])
    for c in ws[1]:
        c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = border

    header_texts = {t['x']: t['text'] for t in tbl['header']}
    for r_idx, row in enumerate(tbl['rows']):
        row_data = {col: '' for col in COLUMN_NAMES}
        for t in row:
            best_col = None
            best_dist = 1e9
            for col in COLUMN_NAMES:
                for hx, htext in header_texts.items():
                    if htext == col:
                        dist = abs(t['x'] - hx)
                        if dist < best_dist and dist < 50:
                            best_dist = dist
                            best_col = col
                        break
            if best_col and t['text']:
                if row_data[best_col]:
                    row_data[best_col] += ' ' + t['text']
                else:
                    row_data[best_col] = t['text']

        ws.append([row_data[c] for c in COLUMN_NAMES])
        r = ws.max_row
        for c_idx in range(1, 11):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = border; cell.alignment = data_align

    for col_letter, w in [('A', 25), ('B', 25), ('C', 12), ('D', 12), ('E', 8),
                          ('F', 8), ('G', 8), ('H', 25), ('I', 15), ('J', 8)]:
        ws.column_dimensions[col_letter].width = w
    ws.freeze_panes = "A2"

wb.save(OUTPUT_XLSX)
print(f"\n✅ Excel 已保存: {OUTPUT_XLSX}")
print(f"   大小: {OUTPUT_XLSX.stat().st_size/1024:.1f} KB")
print(f"\n   Sheet 列表:")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"     {name}: {ws.max_row-1} 行 x {ws.max_column} 列")
