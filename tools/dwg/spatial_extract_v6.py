"""
完整提取多图框表格 - v6 最终版
策略：
- 每个"件号"是一个表格的起点
- 同一表格的标题文字 X 与"件号"X 差 < 250mm
- 数据行：标题 Y 以下 30-300mm 范围，X 与列中心差 < 20mm
"""
import ezdxf
from pathlib import Path
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DXF_PATH = Path(r'E:\KPSNC模具资料\GEELY_BHE20_ICE\2026-04-08_25-8715_Geely BHE20-ICE\1.2D3DMOLd\2D\COOLING\25-8715_点冷却器（已打印）.dxf')
OUTPUT_XLSX = DXF_PATH.parent / "DXF表格提取.xlsx"

def clean_mtext(text: str) -> str:
    text = re.sub(r'\{[^}]*\}', '', text)
    text = re.sub(r'\\[WwFfTtCcPpQqAaLlKk][^;]*;', '', text)
    text = re.sub(r'\\[WwFfTtCcPpQqAaLlKk]', '', text)
    text = text.replace('%%c', 'Ø')
    text = re.sub(r'\s+', ' ', text).strip()
    return text

print(f"读取: {DXF_PATH.name}")
doc = ezdxf.readfile(str(DXF_PATH))
msp = doc.modelspace()
entities = list(msp)

# 收集文字
text_items = []
for e in entities:
    if e.dxftype() in ('TEXT', 'MTEXT'):
        try:
            raw = e.dxf.text if e.dxftype() == 'TEXT' else e.text
            if not raw or not raw.strip():
                continue
            clean = clean_mtext(raw)
            if not clean:
                continue
            pos = e.dxf.insert
            text_items.append({
                'x': float(pos[0]), 'y': float(pos[1]),
                'text': clean, 'raw': raw, 'layer': e.dxf.layer,
            })
        except:
            pass
print(f"文字总数: {len(text_items)}")

# ===== 找所有"件号"位置 (AM_4) =====
anchors = []
for t in text_items:
    if t['text'] == '件号' and t['layer'] == 'AM_4':
        anchors.append({'x': t['x'], 'y': t['y']})

# 按 X 排序
anchors.sort(key=lambda a: a['x'])
print(f"\n件号锚点: {len(anchors)} 个")
for a in anchors:
    print(f"  ({a['x']:.1f}, {a['y']:.1f})")

COLUMN_NAMES = ['件号', '冷却编号', '型号', 'd', 'L', 'E', 'J', '使用零件', '运水孔K', '数量']

# ===== 对每个锚点，提取其表格 =====
# 表格 X 范围：anchor_x - 50 ~ anchor_x + 250
# 表格 Y 范围：header_y - 5 ~ header_y - 200

tables = []
for a in anchors:
    a_x, a_y = a['x'], a['y']

    # 找标题行：a_x ± 300 范围内，Y 接近 a_y (容差 2mm) 的 AM_4 文字
    header_texts = []
    for t in text_items:
        if abs(t['x'] - a_x) < 300 and abs(t['y'] - a_y) < 2.0 and t['layer'] == 'AM_4':
            if t['text'] in COLUMN_NAMES:
                header_texts.append(t)
    header_texts.sort(key=lambda t: t['x'])

    if len(header_texts) < len(COLUMN_NAMES):
        # 标题不完整，跳过
        continue

    # 列 X 中心（按 X 顺序对应 COLUMN_NAMES）
    col_x_map = {}
    for i, t in enumerate(header_texts):
        if i < len(COLUMN_NAMES):
            col_x_map[COLUMN_NAMES[i]] = t['x']

    # 数据行：a_x ± 300, Y = a_y - 5 ~ a_y - 250
    data_texts = [t for t in text_items
                  if abs(t['x'] - a_x) < 300
                  and a_y - 250 <= t['y'] <= a_y - 3
                  and t['layer'] == 'AM_4'
                  and t['text'] not in COLUMN_NAMES]

    # 按 Y 聚类成行
    sorted_data = sorted(data_texts, key=lambda t: -t['y'])
    rows = []
    if sorted_data:
        current = [sorted_data[0]]
        for t in sorted_data[1:]:
            if abs(t['y'] - current[-1]['y']) < 3.0:
                current.append(t)
            else:
                current.sort(key=lambda t: t['x'])
                rows.append(current)
                current = [t]
        current.sort(key=lambda t: t['x'])
        rows.append(current)

    tables.append({
        'anchor_x': a_x, 'anchor_y': a_y,
        'col_x_map': col_x_map,
        'data_rows': rows,
    })

print(f"\n共识别 {len(tables)} 个表格")
for i, t in enumerate(tables):
    print(f"  表格 {i+1}: 锚点 ({t['anchor_x']:.0f}, {t['anchor_y']:.0f}), 数据 {len(t['data_rows'])} 行")

# ===== 输出 Excel =====
wb = Workbook()
wb.remove(wb.active)

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="305496")
thin = Side(border_style="thin", color="999999")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
data_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center")

def assign_to_col(t_x, col_x_map, threshold=20):
    best_col = None
    best_dist = threshold
    for col, x in col_x_map.items():
        d = abs(t_x - x)
        if d < best_dist:
            best_dist = d
            best_col = col
    return best_col

# Sheet 1: 全部表格汇总
ws = wb.create_sheet("全部表格汇总")
ws.append(['表#', '锚点 X', '锚点 Y', '行号'] + COLUMN_NAMES)
for c in ws[1]:
    c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = border

for t_idx, tbl in enumerate(tables):
    for r_idx, row in enumerate(tbl['data_rows']):
        row_data = {c: '' for c in COLUMN_NAMES}
        for t in row:
            col = assign_to_col(t['x'], tbl['col_x_map'])
            if col and t['text']:
                if row_data[col]:
                    row_data[col] += ' ' + t['text']
                else:
                    row_data[col] = t['text']
        ws.append([t_idx + 1, round(tbl['anchor_x']), round(tbl['anchor_y']), r_idx + 1] +
                  [row_data[c] for c in COLUMN_NAMES])
        r = ws.max_row
        for c_idx in range(1, 15):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = border; cell.alignment = data_align

for col, w in [('A', 6), ('B', 10), ('C', 10), ('D', 6),
               ('E', 25), ('F', 25), ('G', 15), ('H', 12), ('I', 8),
               ('J', 8), ('K', 8), ('L', 25), ('M', 15), ('N', 8)]:
    ws.column_dimensions[col].width = w
ws.freeze_panes = "E2"
ws.auto_filter.ref = ws.dimensions

# 每个图框单独 Sheet
for t_idx, tbl in enumerate(tables):
    sheet_name = f"表{t_idx+1}"
    if len(sheet_name) > 31:
        sheet_name = f"T{t_idx+1}"
    ws = wb.create_sheet(sheet_name)
    ws.append(COLUMN_NAMES)
    for c in ws[1]:
        c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = border

    for row in tbl['data_rows']:
        row_data = {c: '' for c in COLUMN_NAMES}
        for t in row:
            col = assign_to_col(t['x'], tbl['col_x_map'])
            if col and t['text']:
                if row_data[col]:
                    row_data[col] += ' ' + t['text']
                else:
                    row_data[col] = t['text']
        ws.append([row_data[c] for c in COLUMN_NAMES])
        r = ws.max_row
        for c_idx in range(1, 11):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = border; cell.alignment = data_align

    for col, w in [('A', 25), ('B', 25), ('C', 15), ('D', 12), ('E', 8),
                  ('F', 8), ('G', 8), ('H', 25), ('I', 15), ('J', 8)]:
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

wb.save(OUTPUT_XLSX)
print(f"\n✅ Excel 已保存: {OUTPUT_XLSX}")
print(f"   大小: {OUTPUT_XLSX.stat().st_size/1024:.1f} KB")
print(f"\nSheet 列表:")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"  {name}: {ws.max_row-1} 行 x {ws.max_column} 列")
