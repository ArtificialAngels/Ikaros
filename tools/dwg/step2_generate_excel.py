"""
江铃模具文件清单 - 生成 Excel
输入：E:\\KPSNC模具资料\\...\\2D 目录
输出：E:\\KPSNC模具资料\\...\\2D\\模具文件清单.xlsx

文件名解析规则：
  1. 标准模式: 25-8777_XXXX_中文描述_编号.dwg  -> 提取中文描述 + 末尾编号
  2. L250864 模式: L250864-C1411 (25-8777)中文描述.dwg -> 提取中文描述
  3. asm 模式: 中文描述.dwg（无前缀） -> 直接保留
  4. 25-8763_S7散件(已打印).dwl -> 特殊处理
"""
import re
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(r"E:\KPSNC模具资料\江铃\江铃缸体GE266 2.3 AN\ge266_2.3an(tac2)压铸模具2D,3D图档数据20250929\ge266_2.3an(tac2)压铸模具2D,3D图档数据20250929\2D")
OUTPUT = ROOT / "模具文件清单.xlsx"

# 项目号前缀集合（需要去除的前缀字样）
PROJECT_PREFIXES = [
    r"25-8777",        # 主项目号
    r"25-8763",        # 散件
    r"225-8777",       # 偶尔出现的变体（极少见）
    r"L250864-C1411\s*\(25-8777\)",  # 模架图纸特殊格式
]

# 拼装正则
PREFIX_PATTERN = re.compile(
    r"^(" + "|".join(PROJECT_PREFIXES) + r")[\s_\-]*",
    re.IGNORECASE
)

# 编号后缀模式：文件名末尾的 _XXXX（4位或5位数字；3位太容易撞到型号代码，5位更稳）
# 注意：传入的已经是 stem（去后缀），所以不需要 lookahead
NUM_SUFFIX_PATTERN = re.compile(r"_(\d{4,5})$", re.IGNORECASE)
# 中文开头匹配（用于没有编号的简单情况）
HAS_CHINESE = re.compile(r"[\u4e00-\u9fff]")

# 临时文件/系统文件后缀（仍保留，但可在备注中标出）
TEMP_EXTS = {".dwl", ".dwl2", ".bak"}


def parse_filename(stem: str) -> dict:
    """
    解析文件名（不含扩展名），返回:
      - prefix_removed: 去除前缀后的字符串
      - description: 中文描述
      - number: 第一组编号
    """
    original = stem
    # 1) 去前缀
    s = PREFIX_PATTERN.sub("", stem, count=1)
    prefix_removed = s

    # 2) 提取末尾编号（_XXXX，4-5 位）
    m = NUM_SUFFIX_PATTERN.search(s)
    number = ""
    desc_part = s
    if m:
        number = m.group(1)
        idx = m.start()
        desc_part = s[:idx]
    else:
        # 备选：找一个形如 _NNNN 的片段作为编号（中文描述前的最后一个数字段）
        # 例：M20吊环孔堵塞_2519 -> 编号 2519
        any_num = re.search(r"_(\d{4,5})\b", s)
        if any_num:
            number = any_num.group(1)
            desc_part = s[:any_num.start()]

    # 清理描述
    desc_part = re.sub(r"[\s_\-]+$", "", desc_part)

    # 取最后一个含中文的 _ 分段作为最终描述
    if HAS_CHINESE.search(desc_part):
        parts = desc_part.split("_")
        for i in range(len(parts) - 1, -1, -1):
            if HAS_CHINESE.search(parts[i]):
                description = parts[i]
                break
        else:
            description = desc_part
    else:
        description = desc_part

    description = description.strip()
    return {
        "original": original,
        "prefix_removed": prefix_removed,
        "description": description,
        "number": number,
    }


def get_extension(path: Path) -> str:
    """获取小写后缀（不含点）"""
    return path.suffix.lower().lstrip(".")


def main():
    print(f"扫描根目录: {ROOT}")
    if not ROOT.exists():
        raise SystemExit(f"❌ 目录不存在: {ROOT}")

    rows = []
    subdirs = sorted([p for p in ROOT.iterdir() if p.is_dir()])

    print(f"发现 {len(subdirs)} 个子目录")

    total_files = 0
    for sd in subdirs:
        files = sorted([p for p in sd.iterdir() if p.is_file()])
        print(f"  📁 {sd.name}: {len(files)} 个文件")
        for fp in files:
            total_files += 1
            stem = fp.stem
            ext = get_extension(fp)

            parsed = parse_filename(stem)

            # 备注
            remarks = []
            if ext in TEMP_EXTS:
                ext_name_map = {"dwl": "AutoCAD 锁文件", "dwl2": "AutoCAD 锁文件", "bak": "备份文件"}
                remarks.append(ext_name_map.get(ext, "临时文件"))
            if not parsed["number"]:
                remarks.append("无编号")
            if not parsed["description"]:
                remarks.append("无中文描述")

            rows.append({
                "序号": total_files,
                "子目录": sd.name,
                "原文件名": fp.name,
                "中文描述": parsed["description"],
                "编号": parsed["number"],
                "扩展名": ext,
                "备注": "; ".join(remarks),
            })

    print(f"\n共解析 {total_files} 个文件")

    # ============ 写入 Excel ============
    wb = Workbook()
    ws = wb.active
    ws.title = "模具文件清单"

    # 表头
    headers = ["序号", "子目录", "原文件名", "中文描述", "编号", "扩展名", "备注"]
    ws.append(headers)

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="305496")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="999999")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for col_idx, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border

    # 数据
    data_align = Alignment(horizontal="left", vertical="center", wrap_text=False)
    for r in rows:
        ws.append([r["序号"], r["子目录"], r["原文件名"], r["中文描述"], r["编号"], r["扩展名"], r["备注"]])
        row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = data_align
            cell.border = border
        # 临时文件行加灰底
        if r["扩展名"] in {"dwl", "dwl2", "bak"}:
            gray = PatternFill("solid", fgColor="F2F2F2")
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = gray

    # 列宽
    col_widths = {
        1: 6,    # 序号
        2: 28,   # 子目录
        3: 60,   # 原文件名
        4: 40,   # 中文描述
        5: 10,   # 编号
        6: 10,   # 扩展名
        7: 20,   # 备注
    }
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = ws.dimensions

    # ============ 摘要 Sheet ============
    from collections import Counter
    ws2 = wb.create_sheet("子目录统计")
    ws2.append(["子目录", "文件总数", "有编号", "无编号", "占总数%"])
    for col_idx in range(1, 6):
        c = ws2.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = border

    sub_counter = Counter(r["子目录"] for r in rows)
    sub_with_num = Counter(r["子目录"] for r in rows if r["编号"])
    total_n = len(rows)
    for sub, cnt in sub_counter.most_common():
        with_n = sub_with_num[sub]
        no_n = cnt - with_n
        pct = cnt * 100 / total_n
        ws2.append([sub, cnt, with_n, no_n, f"{pct:.1f}%"])
        for col_idx in range(1, 6):
            cell = ws2.cell(row=ws2.max_row, column=col_idx)
            cell.border = border
            cell.alignment = data_align

    # 合计行
    ws2.append([])
    ws2.append(["合计", total_n, sum(sub_with_num.values()), total_n - sum(sub_with_num.values()), "100.0%"])
    for col_idx in range(1, 6):
        cell = ws2.cell(row=ws2.max_row, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.border = border

    # 摘要列宽
    for col, w in [(1, 30), (2, 12), (3, 10), (4, 10), (5, 12)]:
        ws2.column_dimensions[get_column_letter(col)].width = w

    wb.save(OUTPUT)
    print(f"\n✅ Excel 已保存: {OUTPUT}")
    print(f"   文件大小: {OUTPUT.stat().st_size / 1024:.1f} KB")

    # 抽检前 10 行打印
    print("\n抽检前 10 行:")
    print("-" * 100)
    for r in rows[:10]:
        print(f"  [{r['序号']:>3}] {r['子目录']:<20} | {r['中文描述']:<35} | 编号: {r['编号']:<6} | {r['原文件名']}")


if __name__ == "__main__":
    main()
