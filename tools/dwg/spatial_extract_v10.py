"""
完整提取多图框表格 - v10 终极版
按用户方法：件号 -> 上方最近横线 + 左侧最近竖线 = 边界左上角 -> 找右下角
"""
import ezdxf
from pathlib import Path
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DXF_PATH = Path(r'E:\KPSNC模具资料\GEELY_BHE20_ICE\2026-04-08_25-8715_Geely BHE20-ICE\1.2D3DMOLd\2D\COOLING\25-8715_点冷却器（已打印）.dxf')
OUTPUT_XLSX = DXF_PATH.parent / "DXF表格提取.xlsx"

def clean_mtext(text):
    text = re.sub(r'\{[^}]*\}', '', text)
    text = re.sub(r'\\[WwFfTtCcPpQqAaLlKk][^;]*;', '', text)
    text = re.sub(r'\\[WwFfTtCcPpQqAaLlKk]', '', text)
    text = text.replace('%%c', 'Ø')
    return re.sub(r'\s+', ' ', text).strip()

print(f"读取: {DXF_PATH.name}")
doc = ezdxf.readfile(str(DXF_PATH))
msp = doc.modelspace()
entities = list(msp)

# ===== 收集 LINE =====
h_lines = []
v_lines = []
for e in entities:
    if e.dxftype() == 'LINE':
        try:
            s = e.dxf.start
            end = e.dxf.end
            x1, y1 = float(s[0]), float(s[1])
            x2, y2 = float(end[0]), float(end[1])
            if x1 > x2:
                x1, x2 = x2, x1
            if y1 > y2:
                y1, y2 = y2, y1
            if abs(x2 - x1) > abs(y2 - y1):
                h_lines.append({'x1': x1, 'y1': y1, 'x2': x2, 'y2': y2})
            else:
                v_lines.append({'x1': x1, 'y1': y1, 'x2': x2, 'y2': y2})
        except:
            pass

# ===== 收集文字 =====
text_items = []
for e in entities:
    if e.dxftype() in ('TEXT', 'MTEXT'):
        try:
            raw = e.dxf.text if e.dxftype() == 'TEXT' else e.text
            if not raw or not raw.strip():
                continue
            clean = clean_mtext(raw)
            if not clean:
                continue
            pos = e.dxf.insert
            text_items.append({
                'x': float(pos[0]), 'y': float(pos[1]),
                'text': clean, 'layer': e.dxf.layer,
            })
        except:
            pass

# ===== 找"件号"位置 =====
jianhao_list = [t for t in text_items if t['text'] == '件号']

# ===== 找"标题行"（Y 簇含 5 关键词且都在 AM_4） =====
COLUMN_NAMES = ['件号', '冷却编号', '型号', 'd', 'L', 'E', 'J', '使用零件', '运水孔K', '数量']
HEADER_KEYS = {'件号', '冷却编号', '型号', '使用零件', '数量'}

sorted_by_y = sorted(text_items, key=lambda t: -t['y'])
y_groups = []
current = [sorted_by_y[0]]
for t in sorted_by_y[1:]:
    if abs(t['y'] - current[-1]['y']) < 2.0:
        current.append(t)
    else:
        y_groups.append(current)
        current = [t]
y_groups.append(current)

header_y_set = set()
for grp in y_groups:
    texts = [t['text'] for t in grp]
    if all(k in texts for k in HEADER_KEYS) and all(t['layer'] == 'AM_4' for t in grp):
        header_y_set.add(round(grp[0]['y'], 2))

# 找标题行：精确等于 (不聚类)
header_texts = [t for t in text_items if t['text'] in COLUMN_NAMES and t['layer'] == 'AM_4']
# 按 (y, x) 排序
header_texts.sort(key=lambda t: (round(t['y'], 2), t['x']))
print(f"标题行文字: {len(header_texts)} 个")

# ===== 对每个标题行 Y，找其"列"和"数据" =====
# 关键：标题行是表格第一行，标题行的紧邻上方横线 = y_top
# 标题行的紧邻下方横线 = y_first_data_divider
# 标题行的紧邻左侧竖线 = x_left_first_col
# 表格最右竖线 = x_right
# 表格最下横线 = y_bottom

# 先按 Y 聚类标题行（同一标题行 Y 聚成一个组）
header_groups = []
current = [header_texts[0]]
for t in header_texts[1:]:
    if abs(t['y'] - current[-1]['y']) < 2.0:
        current.append(t)
    else:
        header_groups.append(current)
        current = [t]
header_groups.append(current)

print(f"标题行组: {len(header_groups)}")
for grp in header_groups:
    y = grp[0]['y']
    print(f"  Y={y:.1f}: {len(grp)} 个文字, X=[{grp[0]['x']:.0f}, {grp[-1]['x']:.0f}]")

# ===== 对每个标题行，找表格边界 + 提取数据 =====
def find_boundary(header_grp):
    """
    给定一组标题行文字（同一 Y），找表格边界。
    """
    y_h = header_grp[0]['y']
    # 标题行 X 范围
    x_h_min = min(t['x'] for t in header_grp)
    x_h_max = max(t['x'] for t in header_grp)

    # 1) 标题行上方的最近横线 (y > y_h，X 包含标题行 X 范围)
    above = [l for l in h_lines if l['y1'] > y_h
             and l['x1'] - 5 <= x_h_min and x_h_max <= l['x2'] + 5]
    above.sort(key=lambda l: l['y1'] - y_h)
    if not above:
        return None
    y_top = above[0]['y1']

    # 2) 找最远的下方横线（用作表格底部）
    far_below = [l for l in h_lines if l['y1'] < y_h - 5
                 and l['x1'] - 50 <= x_h_min and x_h_max <= l['x2'] + 50
                 and (l['x2'] - l['x1']) > 50]
    far_below.sort(key=lambda l: l['y1'])  # Y 最大优先（即最靠上 = 最远距离标题行最近）
    if far_below:
        y_bottom = far_below[-1]['y1']
    else:
        y_bottom = y_h - 200

    # 3) 左侧最近竖线 (x < x_h_min) - 找 Y 范围覆盖 (y_h - 5, y_bottom) 的
    left_v = [l for l in v_lines if l['x1'] < x_h_min - 1
              and not (l['y2'] < y_h - 5 or l['y1'] > y_bottom)]
    left_v.sort(key=lambda l: x_h_min - l['x1'])
    if not left_v:
        return None
    x_left = left_v[0]['x1']

    # 4) 右侧最近竖线 (x > x_h_max) - 找 Y 范围覆盖 (y_h - 5, y_bottom) 的
    right_v = [l for l in v_lines if l['x1'] > x_h_max + 1
               and not (l['y2'] < y_h - 5 or l['y1'] > y_bottom)]
    right_v.sort(key=lambda l: l['x1'] - x_h_max)
    if not right_v:
        return None
    x_right = right_v[0]['x1']

    return {
        'x_left': x_left, 'x_right': x_right,
        'y_top': y_top, 'y_bottom': y_bottom,
    }

# 处理每个标题行组
tables = []
for grp in header_groups:
    boundary = find_boundary(grp)
    if not boundary:
        print(f"  ⚠️ Y={grp[0]['y']:.1f}: 边界失败")
        continue

    # 列 X 中心
    col_x_map = {}
    for i, t in enumerate(sorted(grp, key=lambda t: t['x'])):
        if i < len(COLUMN_NAMES):
            col_x_map[COLUMN_NAMES[i]] = t['x']

    y_h = grp[0]['y']
    # 数据行：边界内、Y < y_h - 3
    data = [t for t in text_items
            if boundary['x_left'] - 1 <= t['x'] <= boundary['x_right'] + 1
            and boundary['y_bottom'] <= t['y'] <= y_h - 3
            and t['text'] not in COLUMN_NAMES]
    sorted_data = sorted(data, key=lambda t: -t['y'])
    rows = []
    if sorted_data:
        current = [sorted_data[0]]
        for t in sorted_data[1:]:
            if abs(t['y'] - current[-1]['y']) < 3.0:
                current.append(t)
            else:
                current.sort(key=lambda t: t['x'])
                rows.append(current)
                current = [t]
        current.sort(key=lambda t: t['x'])
        rows.append(current)

    tables.append({
        'header_y': y_h,
        'header_grp': grp,
        'boundary': boundary,
        'col_x_map': col_x_map,
        'data_rows': rows,
    })

print(f"\n=== 共识别 {len(tables)} 个表格 ===")
for i, t in enumerate(tables):
    b = t['boundary']
    print(f"  表 {i+1}: 标题 Y={t['header_y']:.0f}  "
          f"X=[{b['x_left']:.0f},{b['x_right']:.0f}] Y=[{b['y_top']:.0f},{b['y_bottom']:.0f}]  "
          f"数据 {len(t['data_rows'])} 行")

# ===== 输出 Excel =====
wb = Workbook()
wb.remove(wb.active)

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="305496")
thin = Side(border_style="thin", color="999999")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
data_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center")

def assign_to_col(t_x, col_x_map, threshold=30):
    best_col = None
    best_dist = threshold
    for col, x in col_x_map.items():
        d = abs(t_x - x)
        if d < best_dist:
            best_dist = d
            best_col = col
    return best_col

# Sheet 1: 全部表格汇总
ws = wb.create_sheet("全部表格汇总")
ws.append(['表#', '标题 Y', 'X范围', 'Y范围', '行号'] + COLUMN_NAMES)
for c in ws[1]:
    c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = border

for t_idx, tbl in enumerate(tables):
    b = tbl['boundary']
    for r_idx, row in enumerate(tbl['data_rows']):
        row_data = {c: '' for c in COLUMN_NAMES}
        for t in row:
            col = assign_to_col(t['x'], tbl['col_x_map'])
            if col and t['text']:
                if row_data[col]:
                    row_data[col] += ' ' + t['text']
                else:
                    row_data[col] = t['text']
        ws.append([t_idx + 1, round(tbl['header_y']),
                   f"{b['x_left']:.0f}~{b['x_right']:.0f}",
                   f"{b['y_top']:.0f}~{b['y_bottom']:.0f}",
                   r_idx + 1] + [row_data[c] for c in COLUMN_NAMES])
        r = ws.max_row
        for c_idx in range(1, 16):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = border; cell.alignment = data_align

for col, w in [('A', 6), ('B', 10), ('C', 16), ('D', 16), ('E', 6),
               ('F', 25), ('G', 25), ('H', 15), ('I', 12), ('J', 8),
               ('K', 8), ('L', 8), ('M', 25), ('N', 15), ('O', 8)]:
    ws.column_dimensions[col].width = w
ws.freeze_panes = "F2"
ws.auto_filter.ref = ws.dimensions

# 每个表格单独 Sheet
for t_idx, tbl in enumerate(tables):
    sheet_name = f"表{t_idx+1}"
    if len(sheet_name) > 31:
        sheet_name = f"T{t_idx+1}"
    ws = wb.create_sheet(sheet_name)
    ws.append(COLUMN_NAMES)
    for c in ws[1]:
        c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = border

    for row in tbl['data_rows']:
        row_data = {c: '' for c in COLUMN_NAMES}
        for t in row:
            col = assign_to_col(t['x'], tbl['col_x_map'])
            if col and t['text']:
                if row_data[col]:
                    row_data[col] += ' ' + t['text']
                else:
                    row_data[col] = t['text']
        ws.append([row_data[c] for c in COLUMN_NAMES])
        r = ws.max_row
        for c_idx in range(1, 11):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = border; cell.alignment = data_align

    for col, w in [('A', 25), ('B', 25), ('C', 15), ('D', 12), ('E', 8),
                  ('F', 8), ('G', 8), ('H', 25), ('I', 15), ('J', 8)]:
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

wb.save(OUTPUT_XLSX)
print(f"\n✅ Excel 已保存: {OUTPUT_XLSX}")
print(f"   大小: {OUTPUT_XLSX.stat().st_size/1024:.1f} KB")
print(f"\nSheet 列表:")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"  {name}: {ws.max_row-1} 行 x {ws.max_column} 列")
