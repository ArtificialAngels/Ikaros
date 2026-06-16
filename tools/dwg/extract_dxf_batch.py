"""
批量提取 DXF 表格 - v14 (clean_mtext 修正版)
"""
import ezdxf
from pathlib import Path
import re
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def clean_mtext(text: str) -> str:
    """
    清理 MTEXT 格式控制符 - v14 修正版
    MTEXT 结构:
      {...} 块内: \\Xvalue; 形式是控制参数，最后一个非控制段是显示文本
      块外: \\Xvalue; 形式是控制参数
    """
    # MTEXT 已知控制符（注意：H 不在内，会和 HNC- 冲突）
    CONTROL_CHARS = set('WwFfTtCcPpQqAaLlKkSsUu~')

    def is_control_segment(s: str) -> bool:
        """判断是否控制段（保守：只过滤明显的控制段）"""
        s = s.strip()
        if not s:
            return True
        # 以 \ 开头（控制符前缀）
        if s[0] == '\\':
            return True
        # 以单/双控制字符开头（\T, \W, \f, \C 等）
        if s[0] in CONTROL_CHARS and len(s) <= 2:
            return True
        # 纯单控制字符（不跟其他字符）
        if len(s) == 1 and s[0] in CONTROL_CHARS:
            return True
        # key=value 形式（如 b0, i0, c134, p2）
        if re.match(r'^[a-zA-Z][a-zA-Z0-9]*=', s):
            return True
        # \f字体,|字体,|... 形式（不含中文）
        if '|' in s and not re.search(r'[\u4e00-\u9fff]', s):
            if re.match(r'^[\w\.\-]+(\|[\w\.\-]+)*$', s):
                return True
        return False

    result = []
    i = 0
    while i < len(text):
        if text[i] == '{':
            j = text.find('}', i)
            if j == -1:
                break
            block = text[i+1:j]
            parts = block.split(';')
            last = ''
            for p in reversed(parts):
                if not is_control_segment(p):
                    last = p.strip()
                    break
            result.append(last)
            i = j + 1
        else:
            result.append(text[i])
            i += 1
    s = ''.join(result)
    s = re.sub(r'\\[WwFfTtCcPpQqAaLlKkSsUu~][^;]*;', '', s)
    s = re.sub(r'\\[WwFfTtCcPpQqAaLlKkSsUu~]', '', s)
    s = s.replace('%%c', 'Ø')
    s = re.sub(r'\\P', '\n', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def extract_tables_from_dxf(dxf_path: Path, output_xlsx: Path):
    """从单个 DXF 提取所有表格到 Excel"""
    print(f"  处理: {dxf_path.name}")
    try:
        doc = ezdxf.readfile(str(dxf_path))
    except Exception as e:
        print(f"    ❌ 打开失败: {e}")
        return False

    msp = doc.modelspace()
    entities = list(msp)

    # 收集 LINE
    h_lines, v_lines = [], []
    for e in entities:
        if e.dxftype() == 'LINE':
            try:
                s = e.dxf.start
                end = e.dxf.end
                x1, y1 = float(s[0]), float(s[1])
                x2, y2 = float(end[0]), float(end[1])
                if x1 > x2: x1, x2 = x2, x1
                if y1 > y2: y1, y2 = y2, y1
                if abs(x2 - x1) > abs(y2 - y1):
                    h_lines.append({'x1': x1, 'y1': y1, 'x2': x2, 'y2': y2})
                else:
                    v_lines.append({'x1': x1, 'y1': y1, 'x2': x2, 'y2': y2})
            except:
                pass

    # 收集文字
    text_items = []
    for e in entities:
        if e.dxftype() in ('TEXT', 'MTEXT'):
            try:
                raw = e.dxf.text if e.dxftype() == 'TEXT' else e.text
                if not raw or not raw.strip():
                    continue
                clean = clean_mtext(raw)
                if not clean:
                    continue
                pos = e.dxf.insert
                text_items.append({
                    'x': float(pos[0]), 'y': float(pos[1]),
                    'text': clean, 'layer': e.dxf.layer,
                })
            except:
                pass

    # 标题行
    COLUMN_NAMES = ['件号', '冷却编号', '型号', 'd', 'L', 'E', 'J', '使用零件', '运水孔K', '数量']
    header_texts = [t for t in text_items if t['text'] in COLUMN_NAMES and t['layer'] == 'AM_4']
    header_texts.sort(key=lambda t: (round(t['y'], 1), t['x']))

    y_groups = []
    if header_texts:
        current = [header_texts[0]]
        for t in header_texts[1:]:
            if abs(t['y'] - current[-1]['y']) < 2.0:
                current.append(t)
            else:
                y_groups.append(current)
                current = [t]
        y_groups.append(current)

    def split_by_x_gap(grp, gap_threshold=100):
        if len(grp) < 2:
            return [grp]
        grp_sorted = sorted(grp, key=lambda t: t['x'])
        result = [[grp_sorted[0]]]
        for t in grp_sorted[1:]:
            if t['x'] - result[-1][-1]['x'] > gap_threshold:
                result.append([t])
            else:
                result[-1].append(t)
        return result

    all_sub_groups = []
    for grp in y_groups:
        subs = split_by_x_gap(grp, gap_threshold=100)
        for sub in subs:
            if len(sub) >= 9:
                all_sub_groups.append(sub)

    def find_boundary(header_sub):
        y_h = header_sub[0]['y']
        x_h_min = min(t['x'] for t in header_sub)
        x_h_max = max(t['x'] for t in header_sub)

        above = [l for l in h_lines if l['y1'] > y_h
                 and l['x1'] - 5 <= x_h_min and x_h_max <= l['x2'] + 5]
        above.sort(key=lambda l: l['y1'] - y_h)
        if not above:
            return None
        y_top = above[0]['y1']

        # 下方最近横线（不只是最远）
        # y_h 下方 1mm 内开始的横线就是"数据行 1"（6500-1 在 Y=-4599）
        # 数据行下方 -4608 是"数据行结束横线"
        below = [l for l in h_lines if l['y1'] < y_h - 1
                 and l['x1'] - 5 <= x_h_min and x_h_max <= l['x2'] + 5
                 and l['y1'] > y_h - 50]  # 限制在标题行下方 50mm 内
        # 按 Y 升序：最下方（Y 最小）到最上方（Y 最大）
        below.sort(key=lambda l: l['y1'])
        if below:
            y_bottom = below[0]['y1']  # Y 最小 = 最靠下 = 表格底部
        else:
            y_bottom = y_h - 30

        left_v = [l for l in v_lines if l['x1'] < x_h_min - 1
                  and not (l['y2'] < y_h - 5 or l['y1'] > y_bottom)]
        left_v.sort(key=lambda l: x_h_min - l['x1'])
        if not left_v:
            return None
        x_left = left_v[0]['x1']

        right_v = [l for l in v_lines if l['x1'] > x_h_max + 1
                   and not (l['y2'] < y_h - 5 or l['y1'] > y_bottom)]
        right_v.sort(key=lambda l: l['x1'] - x_h_max)
        if not right_v:
            return None
        x_right = right_v[0]['x1']

        return {
            'x_left': x_left, 'x_right': x_right,
            'y_top': y_top, 'y_bottom': y_bottom,
        }

    tables = []
    for sub in all_sub_groups:
        boundary = find_boundary(sub)
        if not boundary:
            continue

        col_x_map = {}
        sorted_sub = sorted(sub, key=lambda t: t['x'])
        for i, t in enumerate(sorted_sub):
            if i < len(COLUMN_NAMES):
                col_x_map[COLUMN_NAMES[i]] = t['x']

        y_h = sub[0]['y']
        data = [t for t in text_items
                if boundary['x_left'] - 1 <= t['x'] <= boundary['x_right'] + 1
                and boundary['y_bottom'] <= t['y'] <= y_h - 3
                and t['text'] not in COLUMN_NAMES]
        sorted_data = sorted(data, key=lambda t: -t['y'])
        rows = []
        if sorted_data:
            # 自适应 Y 容差：先扫一次找最常见 Y 差，最大取 8mm
            if len(sorted_data) > 1:
                y_diffs = []
                prev_y = sorted_data[0]['y']
                for t in sorted_data[1:]:
                    d = abs(prev_y - t['y'])
                    if d > 1.5:  # 忽略同 Y 内的微小差
                        y_diffs.append(d)
                        prev_y = t['y']
                if y_diffs:
                    # 中位数作为典型行高
                    y_diffs.sort()
                    median_diff = y_diffs[len(y_diffs) // 2]
                    y_tol = min(8.0, max(4.0, median_diff * 0.55))
                else:
                    y_tol = 5.0
            else:
                y_tol = 5.0
            current = [sorted_data[0]]
            for t in sorted_data[1:]:
                # 自适应数据行 Y 容差（基于该表实际行高）
                if abs(t['y'] - current[-1]['y']) < y_tol:
                    current.append(t)
                else:
                    current.sort(key=lambda t: t['x'])
                    rows.append(current)
                    current = [t]
            current.sort(key=lambda t: t['x'])
            rows.append(current)

        tables.append({
            'header_y': y_h,
            'header_sub': sub,
            'boundary': boundary,
            'col_x_map': col_x_map,
            'data_rows': rows,
        })

    total_data_rows = sum(len(t['data_rows']) for t in tables)
    print(f"    → 文字 {len(text_items)} 个, 标题 {len(header_texts)} 个, 识别 {len(tables)} 个表 (共 {total_data_rows} 数据行)")

    if not tables:
        return False

    # 写 Excel
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="305496")
    thin = Side(border_style="thin", color="999999")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    data_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")

    def assign_to_col(t_x, col_x_map, threshold=30):
        best_col = None
        best_dist = threshold
        for col, x in col_x_map.items():
            d = abs(t_x - x)
            if d < best_dist:
                best_dist = d
                best_col = col
        return best_col

    ws = wb.create_sheet("全部表格汇总")
    ws.append(['表#', '标题 Y', 'X范围', 'Y范围', '行号'] + COLUMN_NAMES)
    for c in ws[1]:
        c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = border

    for t_idx, tbl in enumerate(tables):
        b = tbl['boundary']
        for r_idx, row in enumerate(tbl['data_rows']):
            row_data = {c: '' for c in COLUMN_NAMES}
            for t in row:
                col = assign_to_col(t['x'], tbl['col_x_map'])
                if col and t['text']:
                    if row_data[col]:
                        row_data[col] += ' ' + t['text']
                    else:
                        row_data[col] = t['text']
            ws.append([t_idx + 1, round(tbl['header_y']),
                       f"{b['x_left']:.0f}~{b['x_right']:.0f}",
                       f"{b['y_top']:.0f}~{b['y_bottom']:.0f}",
                       r_idx + 1] + [row_data[c] for c in COLUMN_NAMES])
            r = ws.max_row
            for c_idx in range(1, 16):
                cell = ws.cell(row=r, column=c_idx)
                cell.border = border; cell.alignment = data_align

    for col, w in [('A', 6), ('B', 10), ('C', 16), ('D', 16), ('E', 6),
                   ('F', 25), ('G', 25), ('H', 15), ('I', 12), ('J', 8),
                   ('K', 8), ('L', 8), ('M', 25), ('N', 15), ('O', 8)]:
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "F2"
    ws.auto_filter.ref = ws.dimensions

    for t_idx, tbl in enumerate(tables):
        sheet_name = f"表{t_idx+1}"
        if len(sheet_name) > 31:
            sheet_name = f"T{t_idx+1}"
        ws = wb.create_sheet(sheet_name)
        ws.append(COLUMN_NAMES)
        for c in ws[1]:
            c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = border

        for row in tbl['data_rows']:
            row_data = {c: '' for c in COLUMN_NAMES}
            for t in row:
                col = assign_to_col(t['x'], tbl['col_x_map'])
                if col and t['text']:
                    if row_data[col]:
                        row_data[col] += ' ' + t['text']
                    else:
                        row_data[col] = t['text']
            ws.append([row_data[c] for c in COLUMN_NAMES])
            r = ws.max_row
            for c_idx in range(1, 11):
                cell = ws.cell(row=r, column=c_idx)
                cell.border = border; cell.alignment = data_align

        for col, w in [('A', 25), ('B', 25), ('C', 15), ('D', 12), ('E', 8),
                      ('F', 8), ('G', 8), ('H', 25), ('I', 15), ('J', 8)]:
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"

    wb.save(output_xlsx)
    print(f"    ✅ 已保存: {output_xlsx.name}  ({output_xlsx.stat().st_size/1024:.1f} KB)")
    return True


def main():
    if len(sys.argv) < 2:
        print("用法: extract_dxf_batch.py <DXF文件或目录> [输出Excel]")
        sys.exit(1)

    target = Path(sys.argv[1])
    if target.is_file():
        dxf_files = [target]
        if len(sys.argv) >= 3:
            output = Path(sys.argv[2])
        else:
            output = target.parent / f"{target.stem}_表格.xlsx"
    else:
        dxf_files = sorted(target.rglob("*.dxf"))
        output_root = target

    if not dxf_files:
        print("❌ 未找到 DXF 文件")
        sys.exit(1)

    print(f"=== 批量处理 {len(dxf_files)} 个 DXF ===\n")
    success = 0
    for dxf in dxf_files:
        if target.is_dir():
            out = dxf.parent / f"{dxf.stem}_表格.xlsx"
        else:
            out = output
        if extract_tables_from_dxf(dxf, out):
            success += 1

    print(f"\n=== 完成: {success}/{len(dxf_files)} 成功 ===")


if __name__ == "__main__":
    main()
