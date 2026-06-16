"""
完整提取多图框表格 - v9 终极版（按用户方法）
方法：
1. 找"件号"位置 (X0, Y0)
2. 找上方最近的横线（Y 最接近 Y0 但 Y > Y0）
3. 找下方最近的横线（Y 最接近 Y0 但 Y < Y0）
4. 找左侧最近的竖线（X 最接近 X0 但 X < X0） - 在 Y0 附近
5. 找右侧最近的竖线（X 最接近 X0 但 X > X0） - 在 Y0 附近
"""
import ezdxf
from pathlib import Path
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DXF_PATH = Path(r'E:\KPSNC模具资料\GEELY_BHE20_ICE\2026-04-08_25-8715_Geely BHE20-ICE\1.2D3DMOLd\2D\COOLING\25-8715_点冷却器（已打印）.dxf')
OUTPUT_XLSX = DXF_PATH.parent / "DXF表格提取.xlsx"

def clean_mtext(text):
    text = re.sub(r'\{[^}]*\}', '', text)
    text = re.sub(r'\\[WwFfTtCcPpQqAaLlKk][^;]*;', '', text)
    text = re.sub(r'\\[WwFfTtCcPpQqAaLlKk]', '', text)
    text = text.replace('%%c', 'Ø')
    return re.sub(r'\s+', ' ', text).strip()

print(f"读取: {DXF_PATH.name}")
doc = ezdxf.readfile(str(DXF_PATH))
msp = doc.modelspace()
entities = list(msp)

# ===== 收集所有 LINE =====
lines = []
for e in entities:
    if e.dxftype() == 'LINE':
        try:
            s = e.dxf.start
            end = e.dxf.end
            x1, y1 = float(s[0]), float(s[1])
            x2, y2 = float(end[0]), float(end[1])
            if x1 > x2:
                x1, x2 = x2, x1
            if y1 > y2:
                y1, y2 = y2, y1
            is_h = abs(x2 - x1) > abs(y2 - y1)
            if is_h:
                lines.append({'x1': x1, 'y1': y1, 'x2': x2, 'y2': y2,
                              'is_h': True, 'len': x2 - x1})
            else:
                lines.append({'x1': x1, 'y1': y1, 'x2': x2, 'y2': y2,
                              'is_h': False, 'len': y2 - y1})
        except:
            pass

h_lines = [l for l in lines if l['is_h']]
v_lines = [l for l in lines if not l['is_h']]
print(f"LINE: {len(lines)} 总, {len(h_lines)} 水平, {len(v_lines)} 垂直")

# ===== 收集文字 =====
text_items = []
for e in entities:
    if e.dxftype() in ('TEXT', 'MTEXT'):
        try:
            raw = e.dxf.text if e.dxftype() == 'TEXT' else e.text
            if not raw or not raw.strip():
                continue
            clean = clean_mtext(raw)
            if not clean:
                continue
            pos = e.dxf.insert
            text_items.append({
                'x': float(pos[0]), 'y': float(pos[1]),
                'text': clean, 'layer': e.dxf.layer,
            })
        except:
            pass
print(f"文字总数: {len(text_items)}")

# ===== 找"件号"位置 =====
jianhao_list = [t for t in text_items if t['text'] == '件号']
print(f"\n'件号' 锚点: {len(jianhao_list)} 个")

# ===== 找表格边界 =====
def find_table_boundary(x0, y0, h_lines, v_lines):
    """
    从 (x0, y0) 找最近的表格边线：
    - 1) 上方最近横线：Y > y0，Y 最小 (离 y0 最近)
    - 2) 下方最近横线：Y < y0，Y 最大 (离 y0 最近)
    - 3) 左侧最近竖线：X < x0，X 最大 (离 x0 最近)
    - 4) 右侧最近竖线：X > x0，X 最小 (离 x0 最近)
    """
    # 1) 上方最近横线 (Y > y0)
    # 距离 = |Y - y0|，距离最小的 = 最近的
    above = [l for l in h_lines if l['y1'] > y0]
    if not above:
        return None
    above.sort(key=lambda l: l['y1'] - y0)  # 距离 y0 最小
    y_top = above[0]['y1']

    # 2) 下方最近横线 (Y < y0)
    below = [l for l in h_lines if l['y1'] < y0]
    if not below:
        return None
    below.sort(key=lambda l: y0 - l['y1'])  # 距离 y0 最小
    y_bottom = below[0]['y1']

    # 3) 左侧最近竖线 (X < x0)，且该竖线 Y 跨越 (y_top, y_bottom)
    left_v = [l for l in v_lines
              if l['x1'] < x0
              and not (l['y2'] < y_top or l['y1'] > y_bottom)]
    if not left_v:
        return None
    left_v.sort(key=lambda l: x0 - l['x1'])  # 距离 x0 最小
    x_left = left_v[0]['x1']

    # 4) 右侧最近竖线 (X > x0)
    right_v = [l for l in v_lines
               if l['x1'] > x0
               and not (l['y2'] < y_top or l['y1'] > y_bottom)]
    if not right_v:
        return None
    right_v.sort(key=lambda l: l['x1'] - x0)  # 距离 x0 最小
    x_right = right_v[0]['x1']

    return {
        'x_left': x_left, 'x_right': x_right,
        'y_top': y_top, 'y_bottom': y_bottom,
    }

# 处理每个件号
tables = []
COLUMN_NAMES = ['件号', '冷却编号', '型号', 'd', 'L', 'E', 'J', '使用零件', '运水孔K', '数量']

debug_count = 0
for jh in jianhao_list:
    x0, y0 = jh['x'], jh['y']
    boundary = find_table_boundary(x0, y0, h_lines, v_lines)
    print(f"件号 ({x0}, {y0}): boundary={boundary}")
    if not boundary:
        debug_count += 1
        if debug_count <= 3:
            print(f"  → DEBUG #{debug_count} 边界失败")
            above = [l for l in h_lines if l['y1'] > y0]
            below = [l for l in h_lines if l['y1'] < y0]
            above.sort(key=lambda l: l['y1'] - y0)
            below.sort(key=lambda l: y0 - l['y1'])
            if above:
                print(f'    上方最近: Y={above[0]["y1"]:.2f}  X=[{above[0]["x1"]:.0f}, {above[0]["x2"]:.0f}]')
            if below:
                print(f'    下方最近: Y={below[0]["y1"]:.2f}  X=[{below[0]["x1"]:.0f}, {below[0]["x2"]:.0f}]')
        continue

    # 找标题行：边界内、Y 接近 y0 的 5-10 关键词
    header = [t for t in text_items
              if boundary['x_left'] - 1 <= t['x'] <= boundary['x_right'] + 1
              and abs(t['y'] - y0) < 5.0
              and t['text'] in COLUMN_NAMES]
    header.sort(key=lambda t: t['x'])

    if len(header) < len(COLUMN_NAMES):
        # 找最近的"完整标题行" - 尝试扩大 Y 容差
        header = [t for t in text_items
                  if boundary['x_left'] - 1 <= t['x'] <= boundary['x_right'] + 1
                  and abs(t['y'] - y0) < 20.0
                  and t['text'] in COLUMN_NAMES]
        header.sort(key=lambda t: t['x'])

    if len(header) < len(COLUMN_NAMES):
        if debug_count <= 5:
            print(f"  件号 ({x0}, {y0}): 标题不完整 {len(header)}/{len(COLUMN_NAMES)}  边界=({boundary['x_left']:.0f},{boundary['y_top']:.0f})-({boundary['x_right']:.0f},{boundary['y_bottom']:.0f})")
        continue

    # 列 X 中心
    col_x_map = {}
    for i, t in enumerate(header):
        if i < len(COLUMN_NAMES):
            col_x_map[COLUMN_NAMES[i]] = t['x']

    # 数据行：边界内、Y < y0 - 3
    data = [t for t in text_items
            if boundary['x_left'] - 1 <= t['x'] <= boundary['x_right'] + 1
            and boundary['y_bottom'] <= t['y'] <= y0 - 3
            and t['text'] not in COLUMN_NAMES]
    sorted_data = sorted(data, key=lambda t: -t['y'])
    rows = []
    if sorted_data:
        current = [sorted_data[0]]
        for t in sorted_data[1:]:
            if abs(t['y'] - current[-1]['y']) < 3.0:
                current.append(t)
            else:
                current.sort(key=lambda t: t['x'])
                rows.append(current)
                current = [t]
        current.sort(key=lambda t: t['x'])
        rows.append(current)

    tables.append({
        'anchor_x': x0, 'anchor_y': y0,
        'boundary': boundary,
        'col_x_map': col_x_map,
        'data_rows': rows,
    })

print(f"\n=== 共识别 {len(tables)} 个表格 ===")
for i, t in enumerate(tables):
    b = t['boundary']
    print(f"  表 {i+1}: 锚点 ({t['anchor_x']:.0f}, {t['anchor_y']:.0f})  "
          f"X=[{b['x_left']:.0f},{b['x_right']:.0f}] Y=[{b['y_top']:.0f},{b['y_bottom']:.0f}]  "
          f"数据 {len(t['data_rows'])} 行")

# ===== 输出 Excel =====
wb = Workbook()
wb.remove(wb.active)

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="305496")
thin = Side(border_style="thin", color="999999")
border = Border(top=thin, left=thin, right=thin, bottom=thin)
data_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center")

def assign_to_col(t_x, col_x_map, threshold=30):
    best_col = None
    best_dist = threshold
    for col, x in col_x_map.items():
        d = abs(t_x - x)
        if d < best_dist:
            best_dist = d
            best_col = col
    return best_col

# Sheet 1: 全部表格汇总
ws = wb.create_sheet("全部表格汇总")
ws.append(['表#', '锚点 X', '锚点 Y', 'X范围', 'Y范围', '行号'] + COLUMN_NAMES)
for c in ws[1]:
    c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = border

for t_idx, tbl in enumerate(tables):
    b = tbl['boundary']
    for r_idx, row in enumerate(tbl['data_rows']):
        row_data = {c: '' for c in COLUMN_NAMES}
        for t in row:
            col = assign_to_col(t['x'], tbl['col_x_map'])
            if col and t['text']:
                if row_data[col]:
                    row_data[col] += ' ' + t['text']
                else:
                    row_data[col] = t['text']
        ws.append([t_idx + 1, round(tbl['anchor_x']), round(tbl['anchor_y']),
                   f"{b['x_left']:.0f}~{b['x_right']:.0f}",
                   f"{b['y_top']:.0f}~{b['y_bottom']:.0f}",
                   r_idx + 1] + [row_data[c] for c in COLUMN_NAMES])
        r = ws.max_row
        for c_idx in range(1, 17):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = border; cell.alignment = data_align

for col, w in [('A', 6), ('B', 10), ('C', 10), ('D', 14), ('E', 14), ('F', 6),
               ('G', 25), ('H', 25), ('I', 15), ('J', 12), ('K', 8),
               ('L', 8), ('M', 8), ('N', 25), ('O', 15), ('P', 8)]:
    ws.column_dimensions[col].width = w
ws.freeze_panes = "G2"
ws.auto_filter.ref = ws.dimensions

# 每个表格单独 Sheet
for t_idx, tbl in enumerate(tables):
    sheet_name = f"表{t_idx+1}"
    if len(sheet_name) > 31:
        sheet_name = f"T{t_idx+1}"
    ws = wb.create_sheet(sheet_name)
    ws.append(COLUMN_NAMES)
    for c in ws[1]:
        c.font = header_font; c.fill = header_fill; c.alignment = center_align; c.border = border

    for row in tbl['data_rows']:
        row_data = {c: '' for c in COLUMN_NAMES}
        for t in row:
            col = assign_to_col(t['x'], tbl['col_x_map'])
            if col and t['text']:
                if row_data[col]:
                    row_data[col] += ' ' + t['text']
                else:
                    row_data[col] = t['text']
        ws.append([row_data[c] for c in COLUMN_NAMES])
        r = ws.max_row
        for c_idx in range(1, 11):
            cell = ws.cell(row=r, column=c_idx)
            cell.border = border; cell.alignment = data_align

    for col, w in [('A', 25), ('B', 25), ('C', 15), ('D', 12), ('E', 8),
                  ('F', 8), ('G', 8), ('H', 25), ('I', 15), ('J', 8)]:
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

wb.save(OUTPUT_XLSX)
print(f"\n✅ Excel 已保存: {OUTPUT_XLSX}")
print(f"   大小: {OUTPUT_XLSX.stat().st_size/1024:.1f} KB")
print(f"\nSheet 列表:")
for name in wb.sheetnames:
    ws = wb[name]
    print(f"  {name}: {ws.max_row-1} 行 x {ws.max_column} 列")
